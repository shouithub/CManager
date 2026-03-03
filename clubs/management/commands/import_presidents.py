import csv
from pathlib import Path

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from clubs.models import Club, Officer, UserProfile

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


HEADER_ALIASES = {
    'club_name': {'club_name', '社团名称', '社团', 'club'},
    'president_name': {'president_name', '社团负责人', '负责人', 'real_name', '姓名'},
    'phone': {'phone', '联系方式', '手机', '手机号', '联系电话'},
    'username': {'username', '用户名', '账号', '登录名'},
    'password': {'password', '密码'},
    'members_count': {'members_count', '社团人数', '人数', 'member_count'},
    'category': {'category', '类别', '分类'},
}


class Command(BaseCommand):
    help = '批量导入社长账号并绑定社团（支持 CSV/XLSX）'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='导入文件路径（.csv / .xlsx）')
        parser.add_argument('--encoding', type=str, default='utf-8-sig', help='CSV 编码，默认 utf-8-sig')
        parser.add_argument('--default-password', type=str, default='Scm@123456', help='未提供 password 列时使用的默认密码')
        parser.add_argument('--create-missing-club', action='store_true', help='当社团不存在时自动创建')
        parser.add_argument('--dry-run', action='store_true', help='预演，不落库')

    def handle(self, *args, **options):
        file_path = Path(options['file_path']).expanduser().resolve()
        if not file_path.exists():
            raise CommandError(f'文件不存在: {file_path}')

        rows = self._load_rows(file_path, options['encoding'])
        if not rows:
            self.stdout.write(self.style.WARNING('导入文件无数据行，已结束。'))
            return

        summary = {
            'total': 0,
            'created_users': 0,
            'updated_users': 0,
            'created_profiles': 0,
            'updated_profiles': 0,
            'bound_clubs': 0,
            'created_clubs': 0,
            'created_officers': 0,
            'updated_officers': 0,
            'skipped': 0,
        }

        errors = []
        for idx, raw_row in enumerate(rows, start=2):
            summary['total'] += 1
            mapped = self._map_row(raw_row)
            row_result = self._process_row(idx, mapped, options)
            for key in summary:
                if key in row_result:
                    summary[key] += row_result[key]
            if row_result.get('error'):
                errors.append(row_result['error'])

        self.stdout.write(self.style.SUCCESS('导入完成：'))
        self.stdout.write(f"- 总行数: {summary['total']}")
        self.stdout.write(f"- 新建用户: {summary['created_users']}")
        self.stdout.write(f"- 更新用户: {summary['updated_users']}")
        self.stdout.write(f"- 新建资料: {summary['created_profiles']}")
        self.stdout.write(f"- 更新资料: {summary['updated_profiles']}")
        self.stdout.write(f"- 绑定社团: {summary['bound_clubs']}")
        self.stdout.write(f"- 新建社团: {summary['created_clubs']}")
        self.stdout.write(f"- 新建干部记录: {summary['created_officers']}")
        self.stdout.write(f"- 更新干部记录: {summary['updated_officers']}")
        self.stdout.write(f"- 跳过: {summary['skipped']}")

        if errors:
            self.stdout.write(self.style.WARNING('部分行处理失败：'))
            for item in errors[:20]:
                self.stdout.write(f"  - {item}")
            if len(errors) > 20:
                self.stdout.write(f"  - ... 其余 {len(errors) - 20} 条省略")

        if options['dry_run']:
            self.stdout.write(self.style.WARNING('当前为 dry-run，未写入数据库。'))

    def _load_rows(self, file_path: Path, encoding: str):
        suffix = file_path.suffix.lower()
        if suffix == '.csv':
            with file_path.open('r', encoding=encoding, newline='') as file_obj:
                reader = csv.DictReader(file_obj)
                return list(reader)

        if suffix == '.xlsx':
            if load_workbook is None:
                raise CommandError('openpyxl 不可用，无法读取 xlsx 文件。')
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active
            if sheet is None:
                return []
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
            output = []
            for row in rows[1:]:
                output.append({headers[i]: ('' if row[i] is None else str(row[i]).strip()) for i in range(len(headers))})
            return output

        raise CommandError('仅支持 .csv 或 .xlsx 文件')

    def _map_row(self, raw_row: dict):
        normalized = {str(k).strip().lower(): ('' if v is None else str(v).strip()) for k, v in raw_row.items()}
        mapped = {}
        for target_key, aliases in HEADER_ALIASES.items():
            value = ''
            for alias in aliases:
                alias_key = alias.strip().lower()
                if alias_key in normalized and normalized[alias_key] != '':
                    value = normalized[alias_key]
                    break
            mapped[target_key] = value
        return mapped

    def _process_row(self, row_num: int, row: dict, options: dict):
        club_name = row.get('club_name', '').strip()
        president_name = row.get('president_name', '').strip()
        phone = row.get('phone', '').strip()
        username = row.get('username', '').strip() or phone
        password = row.get('password', '').strip() or options['default_password']
        members_count = row.get('members_count', '').strip()
        category = row.get('category', '').strip()

        if not club_name or not president_name or not phone or not username:
            return {
                'skipped': 1,
                'error': f'第 {row_num} 行缺少必要字段（社团名称/负责人/联系方式/用户名）',
            }

        try:
            members_count_value = int(float(members_count)) if members_count else 0
        except ValueError:
            members_count_value = 0

        def do_write():
            result = {
                'created_users': 0,
                'updated_users': 0,
                'created_profiles': 0,
                'updated_profiles': 0,
                'bound_clubs': 0,
                'created_clubs': 0,
                'created_officers': 0,
                'updated_officers': 0,
                'skipped': 0,
            }

            user, user_created = User.objects.get_or_create(username=username)
            if user_created:
                user.set_password(password)
                user.first_name = president_name
                user.save()
                result['created_users'] += 1
            else:
                changed = False
                if president_name and user.first_name != president_name:
                    user.first_name = president_name
                    changed = True
                if changed:
                    user.save(update_fields=['first_name'])
                result['updated_users'] += 1

            profile, profile_created = UserProfile.objects.get_or_create(
                user=user,
                defaults={
                    'role': 'president',
                    'status': 'approved',
                    'real_name': president_name,
                    'phone': phone,
                    'wechat': phone,
                    'political_status': 'non_member',
                },
            )

            if profile_created:
                result['created_profiles'] += 1
            else:
                profile_changed = False
                if profile.role != 'president':
                    profile.role = 'president'
                    profile_changed = True
                if profile.status != 'approved':
                    profile.status = 'approved'
                    profile_changed = True
                if president_name and profile.real_name != president_name:
                    profile.real_name = president_name
                    profile_changed = True
                if phone and profile.phone != phone:
                    profile.phone = phone
                    profile_changed = True
                if phone and not profile.wechat:
                    profile.wechat = phone
                    profile_changed = True
                if profile_changed:
                    profile.save()
                result['updated_profiles'] += 1

            club = Club.objects.filter(name=club_name).first()
            if club is None:
                if not options['create_missing_club']:
                    return {
                        'skipped': 1,
                        'error': f'第 {row_num} 行社团不存在：{club_name}（可加 --create-missing-club）',
                    }
                club = Club.objects.create(
                    name=club_name,
                    description=(f'导入创建，类别：{category}' if category else '导入创建'),
                    founded_date=timezone.localdate(),
                    members_count=members_count_value,
                    status='active',
                    president=user,
                )
                result['created_clubs'] += 1
                result['bound_clubs'] += 1
            else:
                changed_fields = []
                if club.president_id != user.id:
                    club.president = user
                    changed_fields.append('president')
                if members_count and club.members_count != members_count_value:
                    club.members_count = members_count_value
                    changed_fields.append('members_count')
                if changed_fields:
                    club.save(update_fields=changed_fields)
                result['bound_clubs'] += 1

            today = timezone.localdate()
            Officer.objects.filter(
                club=club,
                position='president',
                is_current=True,
            ).exclude(user_profile=profile).update(is_current=False, end_date=today)

            officer, officer_created = Officer.objects.get_or_create(
                club=club,
                user_profile=profile,
                position='president',
                defaults={
                    'appointed_date': today,
                    'is_current': True,
                },
            )
            if officer_created:
                result['created_officers'] += 1
            else:
                officer_changed = False
                if not officer.is_current:
                    officer.is_current = True
                    officer_changed = True
                if officer.end_date is not None:
                    officer.end_date = None
                    officer_changed = True
                if officer_changed:
                    officer.save(update_fields=['is_current', 'end_date'])
                result['updated_officers'] += 1

            return result

        try:
            if options['dry_run']:
                with transaction.atomic():
                    result = do_write()
                    transaction.set_rollback(True)
                    return result
            return do_write()
        except Exception as exc:
            return {
                'skipped': 1,
                'error': f'第 {row_num} 行处理失败：{exc}',
            }
