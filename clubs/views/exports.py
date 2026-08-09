"""
导出相关的视图函数
"""
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.parse

from ..models import RoomBooking, Room, TimeSlot, FormSubmission, FormChannel, PublishedActivity
from ..permissions import has_any_role
from django.utils.translation import gettext as _


_CSV_FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r')


def _safe_export_value(value):
    """防止电子表格公式注入：以 = + - @ 等开头的单元格在 Excel 中会被当作公式。"""
    text = '' if value is None else str(value)
    if text.startswith(_CSV_FORMULA_PREFIXES):
        return "'" + text
    return text


@login_required(login_url='clubs:login')
def export_room_bookings_weekly(request):
    """
    导出房间一周的预约日程为 xlsx 表格
    表格以天为列，时间段为行
    """
    # 检查权限
    if not has_any_role(request.user, 'staff', 'admin'):
        messages.error(request, _('您没有权限导出日程安排'))
        return redirect('clubs:room_calendar')
    
    # 获取房间
    room_id = request.GET.get('room_id')
    if room_id:
        room = get_object_or_404(Room, pk=room_id)
    else:
        # 默认使用第一个房间
        room = Room.objects.first()
        if not room:
            messages.error(request, _('系统中没有房间'))
            return redirect('clubs:room_calendar')

    # 获取周开始日期
    week_start_str = request.GET.get('week_start')
    if week_start_str:
        try:
            chosen_date = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, _('无效的日期格式'))
            return redirect('clubs:room_calendar')
        # 无论选择周几，都归一到该周的周一作为周起始
        week_start = chosen_date - timedelta(days=chosen_date.weekday())
    else:
        # 默认为当前周
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
    
    week_end = week_start + timedelta(days=6)
    
    # 时间段使用管理员配置的启用时间段；未配置时回退到默认时段
    configured_slots = list(TimeSlot.objects.filter(is_active=True).order_by('start_time', 'id'))
    if configured_slots:
        time_slots = [
            {
                'start': slot.start_time,
                'end': slot.end_time,
                'label': f'{slot.label} ({slot.start_time.strftime("%H:%M")}-{slot.end_time.strftime("%H:%M")})',
            }
            for slot in configured_slots
        ]
    else:
        time_slots = [
            {'start': time(8, 15), 'end': time(9, 55), 'label': '第1-2节(8:15-9:55)'},
            {'start': time(10, 5), 'end': time(11, 40), 'label': '第3-4节(10:05-11:40)'},
            {'start': time(11, 40), 'end': time(13, 0), 'label': '午休(11:40-13:00)'},
            {'start': time(13, 0), 'end': time(14, 35), 'label': '第5-6节(13:00-14:35)'},
            {'start': time(14, 45), 'end': time(16, 20), 'label': '第7-8节(14:45-16:20)'},
            {'start': time(16, 20), 'end': time(18, 0), 'label': '课外时间(16:20-18:00)'},
            {'start': time(18, 0), 'end': time(19, 0), 'label': '晚餐(18:00-19:00)'},
            {'start': time(19, 0), 'end': time(20, 0), 'label': '晚间1(19:00-20:00)'},
            {'start': time(20, 0), 'end': time(21, 0), 'label': '晚间2(20:00-21:00)'},
            {'start': time(21, 0), 'end': time(22, 0), 'label': '晚间3(21:00-22:00)'},
        ]
    
    # 获取该周的所有有效预约
    bookings = RoomBooking.objects.filter(
        room=room,
        booking_date__gte=week_start,
        booking_date__lte=week_end,
        status='active'
    ).select_related('user__profile', 'club').order_by('booking_date', 'start_time')
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = f"{room.name}日程-{week_start.strftime('%Y年%m月%d日')}"
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    time_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    time_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置列宽
    ws.column_dimensions['A'].width = 22
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # 写入标题
    ws['A1'] = f"{room.name}日程安排 ({week_start.strftime('%Y年%m月%d日')} - {week_end.strftime('%m月%d日')})"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = center_alignment
    
    # 写入日期行
    ws['A2'] = '时间段'
    ws['A2'].fill = header_fill
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment
    ws['A2'].border = border
    
    weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    for i, day_offset in enumerate(range(7)):
        current_date = week_start + timedelta(days=day_offset)
        col = i + 2
        col_letter = get_column_letter(col)
        
        # 写入日期和星期
        header_text = f"{weekdays[i]}\n{current_date.strftime('%m-%d')}"
        ws[f'{col_letter}2'] = header_text
        ws[f'{col_letter}2'].fill = header_fill
        ws[f'{col_letter}2'].font = header_font
        ws[f'{col_letter}2'].alignment = center_alignment
        ws[f'{col_letter}2'].border = border
    
    # 写入时间段和预约信息
    for row_idx, slot in enumerate(time_slots, start=3):
        row = row_idx
        
        # 时间段标签
        ws[f'A{row}'] = slot['label']
        ws[f'A{row}'].fill = time_fill
        ws[f'A{row}'].font = time_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'A{row}'].number_format = '@'
        
        # 为每一天填充预约信息
        for day_offset in range(7):
            current_date = week_start + timedelta(days=day_offset)
            col = day_offset + 2
            col_letter = get_column_letter(col)
            
            # 获取该日期该时间段的预约
            day_bookings = bookings.filter(booking_date=current_date)
            
            # 找出与该时间段有重叠的预约
            slot_bookings = []
            for booking in day_bookings:
                # 检查预约时间是否与时间段有重叠
                if booking.start_time < slot['end'] and booking.end_time > slot['start']:
                    slot_bookings.append(booking)
            
            # 显示预约信息
            if slot_bookings:
                booking_info = []
                for booking in slot_bookings:
                    club_name = booking.club.name if booking.club else "未关联社团"
                    info = f"{club_name}\n({booking.start_time.strftime('%H:%M')}-{booking.end_time.strftime('%H:%M')})\n{booking.purpose}"
                    booking_info.append(info)
                
                cell_value = _safe_export_value('\n---\n'.join(booking_info))
                ws[f'{col_letter}{row}'] = cell_value
                ws[f'{col_letter}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            else:
                ws[f'{col_letter}{row}'] = ''
            
            ws[f'{col_letter}{row}'].alignment = Alignment(
                horizontal="center", 
                vertical="center", 
                wrap_text=True
            )
            ws[f'{col_letter}{row}'].border = border
        
        # 设置行高
        ws.row_dimensions[row].height = 60
    
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    
    # 设置页边距
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 1
    ws.page_margins.bottom = 1
    
    # 设置打印区域和按页适配
    ws.print_area = f'A1:H{2 + len(time_slots)}'
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # 生成响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{room.name}日程-{week_start.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(filename)}'
    
    wb.save(response)
    return response







# ---- Dynamic form exports ------------------------------------------------------

def export_activities(request):
    import csv
    from django.http import HttpResponse
    activities = PublishedActivity.objects.select_related('club')
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=\"activities.csv\"'
    writer = csv.writer(response)
    writer.writerow(['社团', '活动名称', '活动日期', '开始时间', '结束时间', '地点', '公开报名'])
    for item in activities:
        writer.writerow([
            _safe_export_value(item.club.name),
            _safe_export_value(item.activity_name),
            _safe_export_value(item.activity_date),
            _safe_export_value(item.activity_time_start),
            _safe_export_value(item.activity_time_end),
            _safe_export_value(item.activity_location),
            '是' if item.is_public else '否',
        ])
    return response


def export_audit_center_data(request, tab):
    import csv
    from django.http import HttpResponse
    if not request.user.is_authenticated:
        return redirect('clubs:login')
    try:
        if request.user.profile.role not in ['staff', 'admin']:
            return redirect('clubs:index')
    except Exception:
        return redirect('clubs:index')

    slug = tab.replace('_', '-')
    channel = FormChannel.objects.filter(slug=tab).first()
    if channel is None:
        channel = FormChannel.objects.filter(slug=slug).first()
    submissions = (
        FormSubmission.objects.filter(channel=channel)
        .select_related('channel', 'club', 'submitter', 'reviewer')
        .order_by('-submitted_at')
    ) if channel else FormSubmission.objects.none()
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    export_slug = channel.slug if channel else slug
    response['Content-Disposition'] = f'attachment; filename=\"{export_slug}-submissions.csv\"'
    writer = csv.writer(response)
    writer.writerow(['通道', '社团', '标题', '提交人', '状态', '提交时间', '审核人', '审核时间', '审核意见'])
    for item in submissions:
        writer.writerow([
            _safe_export_value(item.channel.name),
            _safe_export_value(item.club.name),
            _safe_export_value(item.display_title),
            _safe_export_value(item.submitter.username),
            _safe_export_value(item.get_status_display()),
            _safe_export_value(item.submitted_at.strftime('%Y-%m-%d %H:%M') if item.submitted_at else ''),
            _safe_export_value(item.reviewer.username if item.reviewer else ''),
            _safe_export_value(item.reviewed_at.strftime('%Y-%m-%d %H:%M') if item.reviewed_at else ''),
            _safe_export_value(item.review_comment),
        ])
    return response
