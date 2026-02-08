"""
导出相关的视图函数
"""
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.parse

from .models import RoomBooking, Room, ActivityApplication
from .views import is_staff_or_admin


def _is_staff(user):
    """检查用户是否为干事"""
    if not user.is_authenticated:
        return False
    try:
        return user.profile.role == 'staff'
    except:
        return False


def _is_admin(user):
    """检查用户是否为管理员"""
    if not user.is_authenticated:
        return False
    try:
        return user.profile.role == 'admin'
    except:
        return False


@login_required(login_url='clubs:login')
def export_room_bookings_weekly(request):
    """
    导出房间一周的预约日程为 xlsx 表格
    表格以天为列，时间段为行
    """
    # 检查权限
    if not is_staff_or_admin(request.user):
        messages.error(request, '您没有权限导出日程安排')
        return redirect('clubs:room_calendar')
    
    # 获取房间
    room_id = request.GET.get('room_id')
    if room_id:
        room = get_object_or_404(Room, pk=room_id)
    else:
        # 默认使用第一个房间
        room = Room.objects.first()
        if not room:
            messages.error(request, '系统中没有房间')
            return redirect('clubs:room_calendar')

    # 获取周开始日期
    week_start_str = request.GET.get('week_start')
    if week_start_str:
        try:
            week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, '无效的日期格式')
            return redirect('clubs:room_calendar')
    else:
        # 默认为当前周
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
    
    week_end = week_start + timedelta(days=6)
    
    # 定义时间段
    time_slots = [
        {'start': time(8, 15), 'end': time(9, 55), 'label': '第1-2节(8:15-9:55)'},
        {'start': time(10, 5), 'end': time(11, 40), 'label': '第3-4节(10:05-11:40)'},
        {'start': time(11, 40), 'end': time(13, 0), 'label': '午休(11:40-13:00)'},
        {'start': time(13, 0), 'end': time(14, 35), 'label': '第5-6节(13:00-14:35)'},
        {'start': time(14, 45), 'end': time(16, 20), 'label': '第7-8节(14:45-16:20)'},
        {'start': time(16, 20), 'end': time(18, 0), 'label': '课外时间(16:20-18:00)'},
        {'start': time(18, 0), 'end': time(19, 0), 'label': '晚餐(18:00-19:00)'},
        {'start': time(19, 0), 'end': time(20, 0), 'label': '晚间1(19:00-20:00)'},
        {'start': time(20, 0), 'end': time(21, 0), 'label': '晚间2(20:00-21:00)'},
        {'start': time(21, 0), 'end': time(22, 0), 'label': '晚间3(21:00-22:00)'},
    ]
    
    # 获取该周的所有有效预约
    bookings = RoomBooking.objects.filter(
        room=room,
        booking_date__gte=week_start,
        booking_date__lte=week_end,
        status='active'
    ).select_related('user__profile', 'club').order_by('booking_date', 'start_time')
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = f"{room.name}日程-{week_start.strftime('%Y年%m月%d日')}"
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    time_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    time_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置列宽
    ws.column_dimensions['A'].width = 22
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # 写入标题
    ws['A1'] = f"{room.name}日程安排 ({week_start.strftime('%Y年%m月%d日')} - {week_end.strftime('%m月%d日')})"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = center_alignment
    
    # 写入日期行
    ws['A2'] = '时间段'
    ws['A2'].fill = header_fill
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment
    ws['A2'].border = border
    
    weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    for i, day_offset in enumerate(range(7)):
        current_date = week_start + timedelta(days=day_offset)
        col = i + 2
        col_letter = get_column_letter(col)
        
        # 写入日期和星期
        header_text = f"{weekdays[i]}\n{current_date.strftime('%m-%d')}"
        ws[f'{col_letter}2'] = header_text
        ws[f'{col_letter}2'].fill = header_fill
        ws[f'{col_letter}2'].font = header_font
        ws[f'{col_letter}2'].alignment = center_alignment
        ws[f'{col_letter}2'].border = border
    
    # 写入时间段和预约信息
    for row_idx, slot in enumerate(time_slots, start=3):
        row = row_idx
        
        # 时间段标签
        ws[f'A{row}'] = slot['label']
        ws[f'A{row}'].fill = time_fill
        ws[f'A{row}'].font = time_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'A{row}'].number_format = '@'
        
        # 为每一天填充预约信息
        for day_offset in range(7):
            current_date = week_start + timedelta(days=day_offset)
            col = day_offset + 2
            col_letter = get_column_letter(col)
            
            # 获取该日期该时间段的预约
            day_bookings = bookings.filter(booking_date=current_date)
            
            # 找出与该时间段有重叠的预约
            slot_bookings = []
            for booking in day_bookings:
                # 检查预约时间是否与时间段有重叠
                if booking.start_time < slot['end'] and booking.end_time > slot['start']:
                    slot_bookings.append(booking)
            
            # 显示预约信息
            if slot_bookings:
                booking_info = []
                for booking in slot_bookings:
                    club_name = booking.club.name if booking.club else "未关联社团"
                    info = f"{club_name}\n({booking.start_time.strftime('%H:%M')}-{booking.end_time.strftime('%H:%M')})\n{booking.purpose}"
                    booking_info.append(info)
                
                cell_value = '\n---\n'.join(booking_info)
                ws[f'{col_letter}{row}'] = cell_value
                ws[f'{col_letter}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            else:
                ws[f'{col_letter}{row}'] = ''
            
            ws[f'{col_letter}{row}'].alignment = Alignment(
                horizontal="center", 
                vertical="center", 
                wrap_text=True
            )
            ws[f'{col_letter}{row}'].border = border
        
        # 设置行高
        ws.row_dimensions[row].height = 60
    
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    
    # 设置页边距
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 1
    ws.page_margins.bottom = 1
    
    # 设置打印区域和按页适配
    ws.print_area = f'A1:H{2 + len(time_slots)}'
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # 生成响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{room.name}日程-{week_start.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(filename)}'
    
    wb.save(response)
    return response


@login_required(login_url='clubs:login')
def export_activities(request):
    """
    导出活动申请为 Excel 文件
    支持筛选条件：社团名称、活动类型、日期、活动ID列表
    管理员、干事可以导出所有批准的活动；社长只能导出本社团的活动
    """
    from .models import Officer
    
    # 检查权限 - 干事、管理员或社长
    user_role = getattr(request.user.profile, 'role', None) if hasattr(request.user, 'profile') else None
    if user_role not in ['staff', 'admin', 'president']:
        messages.error(request, '您没有权限导出活动数据')
        return redirect('clubs:public_activities')
    
    # 获取筛选参数
    club_filter = request.GET.get('club', '')
    activity_type_filter = request.GET.get('activity_type', '')
    date_filter = request.GET.get('date', '')
    search_query = request.GET.get('search', '')
    activity_ids = request.GET.get('ids', '')
    
    # 获取基础查询集
    if activity_ids:
        # 如果指定了活动ID列表，只导出这些活动
        id_list = [int(id.strip()) for id in activity_ids.split(',') if id.strip().isdigit()]
        activities = ActivityApplication.objects.filter(id__in=id_list)
    else:
        # 否则根据筛选条件获取所有已批准的活动
        activities = ActivityApplication.objects.filter(status='approved')
    
    # 根据用户角色过滤活动
    if user_role == 'president':
        # 社长只能导出本社团的活动
        try:
            user_club = Officer.objects.get(user_profile=request.user.profile, position='president', is_current=True).club
            activities = activities.filter(club=user_club)
            # 如果指定了club_filter，验证是否与用户所属社团一致
            if club_filter and club_filter != user_club.name:
                messages.warning(request, f'您只能导出 {user_club.name} 的活动')
                club_filter = user_club.name
        except Officer.DoesNotExist:
            # 如果找不到社长职位，显示空结果
            activities = activities.none()
    
    # 应用筛选条件
    if club_filter:
        activities = activities.filter(club__name__icontains=club_filter)
    
    if activity_type_filter:
        activities = activities.filter(activity_type=activity_type_filter)
    
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            activities = activities.filter(activity_date=filter_date)
        except ValueError:
            pass
    
    if search_query:
        activities = activities.filter(
            Q(club__name__icontains=search_query) |
            Q(activity_name__icontains=search_query)
        )
    
    activities = activities.select_related('club').order_by('activity_date', 'activity_time_start')
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "活动列表"
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 设置列宽
    columns = [
        ('A', '序号', 6),
        ('B', '社团名称', 20),
        ('C', '活动名称', 25),
        ('D', '活动类型', 15),
        ('E', '活动日期', 12),
        ('F', '开始时间', 10),
        ('G', '结束时间', 10),
        ('H', '活动地点', 20),
        ('I', '预计人数', 10),
        ('J', '联系人', 12),
        ('K', '联系电话', 15),
        ('L', '活动简介', 40),
        ('M', '审核状态', 12),
    ]
    
    # 写入表头
    for col_letter, header_text, width in columns:
        ws[f'{col_letter}1'] = header_text
        ws[f'{col_letter}1'].fill = header_fill
        ws[f'{col_letter}1'].font = header_font
        ws[f'{col_letter}1'].alignment = center_alignment
        ws[f'{col_letter}1'].border = border
        ws.column_dimensions[col_letter].width = width
    
    # 写入数据
    for idx, activity in enumerate(activities, start=2):
        # 序号
        ws[f'A{idx}'] = idx - 1
        ws[f'A{idx}'].alignment = center_alignment
        ws[f'A{idx}'].border = border
        
        # 社团名称
        ws[f'B{idx}'] = activity.club.name
        ws[f'B{idx}'].alignment = center_alignment
        ws[f'B{idx}'].border = border
        
        # 活动名称
        ws[f'C{idx}'] = activity.activity_name
        ws[f'C{idx}'].alignment = left_alignment
        ws[f'C{idx}'].border = border
        
        # 活动类型
        ws[f'D{idx}'] = activity.get_activity_type_display()
        ws[f'D{idx}'].alignment = center_alignment
        ws[f'D{idx}'].border = border
        
        # 活动日期
        ws[f'E{idx}'] = activity.activity_date.strftime('%Y-%m-%d') if activity.activity_date else ''
        ws[f'E{idx}'].alignment = center_alignment
        ws[f'E{idx}'].border = border
        
        # 开始时间
        ws[f'F{idx}'] = activity.activity_time_start.strftime('%H:%M') if activity.activity_time_start else ''
        ws[f'F{idx}'].alignment = center_alignment
        ws[f'F{idx}'].border = border
        
        # 结束时间
        ws[f'G{idx}'] = activity.activity_time_end.strftime('%H:%M') if activity.activity_time_end else ''
        ws[f'G{idx}'].alignment = center_alignment
        ws[f'G{idx}'].border = border
        
        # 活动地点
        ws[f'H{idx}'] = activity.activity_location
        ws[f'H{idx}'].alignment = left_alignment
        ws[f'H{idx}'].border = border
        
        # 预计人数
        ws[f'I{idx}'] = activity.expected_participants
        ws[f'I{idx}'].alignment = center_alignment
        ws[f'I{idx}'].border = border
        
        # 联系人
        ws[f'J{idx}'] = activity.contact_person
        ws[f'J{idx}'].alignment = center_alignment
        ws[f'J{idx}'].border = border
        
        # 联系电话
        ws[f'K{idx}'] = activity.contact_phone
        ws[f'K{idx}'].alignment = center_alignment
        ws[f'K{idx}'].border = border
        
        # 活动简介
        ws[f'L{idx}'] = activity.activity_description
        ws[f'L{idx}'].alignment = left_alignment
        ws[f'L{idx}'].border = border
        
        # 审核状态
        status_map = {
            'pending': '待审核',
            'approved': '已通过',
            'rejected': '已驳回',
        }
        ws[f'M{idx}'] = status_map.get(activity.status, activity.status)
        ws[f'M{idx}'].alignment = center_alignment
        ws[f'M{idx}'].border = border
        
        # 设置行高
        ws.row_dimensions[idx].height = 30
    
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    
    # 生成响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # 生成文件名
    filename = f"活动列表-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(filename)}'
    
    wb.save(response)
    return response


@login_required(login_url='clubs:login')
def export_audit_center_data(request, tab):
    """
    导出审核中心数据为Excel，支持筛选
    """
    # 检查权限
    if not _is_staff(request.user) and not _is_admin(request.user):
        messages.error(request, '您没有权限导出数据')
        return redirect('clubs:staff_audit_center', tab=tab)
    
    from .models import (
        ReviewSubmission, ClubRegistration, ClubRegistrationRequest,
        Reimbursement, ActivityApplication, PresidentTransition,
        SubmissionReview, ClubRegistrationReview, ClubApplicationReview,
        ReimbursementHistory
    )
    
    # 获取筛选参数
    club_name = request.GET.get('club_name', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    status = request.GET.get('status', '').strip()
    
    # 将连字符转换为下划线
    tab_internal = tab.replace('-', '_')
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 根据tab类型设置不同的列和数据
    if tab_internal == 'annual_review':
        ws.title = "年审记录"
        headers = ['申请ID', '社团名称', '提交年份', '状态', '提交时间', '提交次数', '审核ID', '审核人', '审核时间', '审核状态', '审核意见']
        
        # 获取数据并应用筛选条件
        items = ReviewSubmission.objects.all()
        if club_name:
            items = items.filter(club__name__icontains=club_name)
        if start_date:
            items = items.filter(submitted_at__gte=start_date)
        if end_date:
            items = items.filter(submitted_at__lte=end_date + ' 23:59:59')
        if status:
            items = items.filter(status=status)
        items = items.order_by('-submitted_at')
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 写入数据 - 包含所有历史审核记录
        row_num = 2
        for item in items:
            # 获取所有审核记录
            reviews = SubmissionReview.objects.filter(submission=item).order_by('submission_attempt', '-reviewed_at')
            
            if reviews.exists():
                for review in reviews:
                    ws.cell(row=row_num, column=1, value=item.id).border = border
                    ws.cell(row=row_num, column=2, value=item.club.name).border = border
                    ws.cell(row=row_num, column=3, value=item.submission_year).border = border
                    ws.cell(row=row_num, column=4, value=item.get_status_display()).border = border
                    ws.cell(row=row_num, column=5, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                    ws.cell(row=row_num, column=6, value=item.resubmission_attempt).border = border
                    ws.cell(row=row_num, column=7, value=review.id).border = border
                    reviewer_name = ''
                    if review.reviewer:
                        try:
                            reviewer_name = review.reviewer.profile.get_full_name() if hasattr(review.reviewer, 'profile') else review.reviewer.username
                        except:
                            reviewer_name = review.reviewer.username
                    ws.cell(row=row_num, column=8, value=reviewer_name).border = border
                    ws.cell(row=row_num, column=9, value=review.reviewed_at.strftime('%Y-%m-%d %H:%M') if review.reviewed_at else '').border = border
                    ws.cell(row=row_num, column=10, value=review.get_status_display()).border = border
                    ws.cell(row=row_num, column=11, value=review.comment or '').border = border
                    row_num += 1
            else:
                # 没有审核记录时也显示申请信息
                ws.cell(row=row_num, column=1, value=item.id).border = border
                ws.cell(row=row_num, column=2, value=item.club.name).border = border
                ws.cell(row=row_num, column=3, value=item.submission_year).border = border
                ws.cell(row=row_num, column=4, value=item.get_status_display()).border = border
                ws.cell(row=row_num, column=5, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                ws.cell(row=row_num, column=6, value=item.resubmission_attempt).border = border
                ws.cell(row=row_num, column=7, value='').border = border
                ws.cell(row=row_num, column=8, value='').border = border
                ws.cell(row=row_num, column=9, value='').border = border
                ws.cell(row=row_num, column=10, value='').border = border
                ws.cell(row=row_num, column=11, value='').border = border
                row_num += 1
        
        filename = f"年审记录-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
    elif tab_internal == 'registration':
        ws.title = "社团注册"
        headers = ['申请ID', '社团名称', '状态', '提交时间', '提交次数', '审核ID', '审核人', '审核时间', '审核状态', '审核意见']
        
        # 获取数据并应用筛选条件
        items = ClubRegistration.objects.all()
        if club_name:
            items = items.filter(club__name__icontains=club_name)
        if start_date:
            items = items.filter(submitted_at__gte=start_date)
        if end_date:
            items = items.filter(submitted_at__lte=end_date + ' 23:59:59')
        if status:
            items = items.filter(status=status)
        items = items.order_by('-submitted_at')
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 写入数据 - 包含所有历史审核记录
        row_num = 2
        for item in items:
            reviews = ClubRegistrationReview.objects.filter(registration=item).order_by('submission_attempt', '-reviewed_at')
            
            if reviews.exists():
                for review in reviews:
                    ws.cell(row=row_num, column=1, value=item.id).border = border
                    ws.cell(row=row_num, column=2, value=item.club.name).border = border
                    ws.cell(row=row_num, column=3, value=item.get_status_display()).border = border
                    ws.cell(row=row_num, column=4, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                    ws.cell(row=row_num, column=5, value=item.resubmission_attempt).border = border
                    ws.cell(row=row_num, column=6, value=review.id).border = border
                    reviewer_name = ''
                    if review.reviewer:
                        try:
                            reviewer_name = review.reviewer.profile.get_full_name() if hasattr(review.reviewer, 'profile') else review.reviewer.username
                        except:
                            reviewer_name = review.reviewer.username
                    ws.cell(row=row_num, column=7, value=reviewer_name).border = border
                    ws.cell(row=row_num, column=8, value=review.reviewed_at.strftime('%Y-%m-%d %H:%M') if review.reviewed_at else '').border = border
                    ws.cell(row=row_num, column=9, value=review.get_status_display()).border = border
                    ws.cell(row=row_num, column=10, value=review.comment or '').border = border
                    row_num += 1
            else:
                ws.cell(row=row_num, column=1, value=item.id).border = border
                ws.cell(row=row_num, column=2, value=item.club.name).border = border
                ws.cell(row=row_num, column=3, value=item.get_status_display()).border = border
                ws.cell(row=row_num, column=4, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                ws.cell(row=row_num, column=5, value=item.resubmission_attempt).border = border
                for col in range(6, 11):
                    ws.cell(row=row_num, column=col, value='').border = border
                row_num += 1
        
        filename = f"社团注册-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
    elif tab_internal == 'application':
        ws.title = "社团申请"
        headers = ['申请ID', '社团名称', '状态', '提交时间', '提交次数', '审核ID', '审核人', '审核时间', '审核状态', '审核意见']
        
        # 获取数据并应用筛选条件
        items = ClubRegistrationRequest.objects.all()
        if club_name:
            items = items.filter(club_name__icontains=club_name)
        if start_date:
            items = items.filter(submitted_at__gte=start_date)
        if end_date:
            items = items.filter(submitted_at__lte=end_date + ' 23:59:59')
        if status:
            items = items.filter(status=status)
        items = items.order_by('-submitted_at')
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 写入数据 - 包含所有历史审核记录
        row_num = 2
        for item in items:
            reviews = ClubApplicationReview.objects.filter(application=item).order_by('submission_attempt', '-reviewed_at')
            
            if reviews.exists():
                for review in reviews:
                    ws.cell(row=row_num, column=1, value=item.id).border = border
                    ws.cell(row=row_num, column=2, value=item.club_name).border = border
                    ws.cell(row=row_num, column=3, value=item.get_status_display()).border = border
                    ws.cell(row=row_num, column=4, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                    ws.cell(row=row_num, column=5, value=item.resubmission_attempt).border = border
                    ws.cell(row=row_num, column=6, value=review.id).border = border
                    reviewer_name = ''
                    if review.reviewer:
                        try:
                            reviewer_name = review.reviewer.profile.get_full_name() if hasattr(review.reviewer, 'profile') else review.reviewer.username
                        except:
                            reviewer_name = review.reviewer.username
                    ws.cell(row=row_num, column=7, value=reviewer_name).border = border
                    ws.cell(row=row_num, column=8, value=review.reviewed_at.strftime('%Y-%m-%d %H:%M') if review.reviewed_at else '').border = border
                    ws.cell(row=row_num, column=9, value=review.get_status_display()).border = border
                    ws.cell(row=row_num, column=10, value=review.comment or '').border = border
                    row_num += 1
            else:
                ws.cell(row=row_num, column=1, value=item.id).border = border
                ws.cell(row=row_num, column=2, value=item.club_name).border = border
                ws.cell(row=row_num, column=3, value=item.get_status_display()).border = border
                ws.cell(row=row_num, column=4, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                ws.cell(row=row_num, column=5, value=item.resubmission_attempt).border = border
                for col in range(6, 11):
                    ws.cell(row=row_num, column=col, value='').border = border
                row_num += 1
        
        filename = f"社团申请-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
    elif tab_internal == 'reimbursement':
        ws.title = "报销记录"
        headers = ['申请ID', '社团名称', '报销日期', '报销金额', '状态', '提交时间', '提交次数', '审核ID', '审核人', '审核时间', '审核状态', '审核意见']
        
        # 获取数据并应用筛选条件
        items = Reimbursement.objects.all()
        if club_name:
            items = items.filter(club__name__icontains=club_name)
        if start_date:
            items = items.filter(submitted_at__gte=start_date)
        if end_date:
            items = items.filter(submitted_at__lte=end_date + ' 23:59:59')
        if status:
            items = items.filter(status=status)
        items = items.order_by('-submitted_at')
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 写入数据 - 包含所有历史审核记录
        row_num = 2
        for item in items:
            reviews = ReimbursementHistory.objects.filter(reimbursement=item).order_by('attempt_number')
            
            if reviews.exists():
                for review in reviews:
                    ws.cell(row=row_num, column=1, value=item.id).border = border
                    ws.cell(row=row_num, column=2, value=item.club.name).border = border
                    ws.cell(row=row_num, column=3, value=item.submission_date.strftime('%Y-%m-%d')).border = border
                    ws.cell(row=row_num, column=4, value=float(item.reimbursement_amount)).border = border
                    ws.cell(row=row_num, column=5, value=item.get_status_display()).border = border
                    ws.cell(row=row_num, column=6, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                    ws.cell(row=row_num, column=7, value=item.resubmission_attempt).border = border
                    ws.cell(row=row_num, column=8, value=review.id).border = border
                    reviewer_name = ''
                    if review.reviewer:
                        try:
                            reviewer_name = review.reviewer.profile.get_full_name() if hasattr(review.reviewer, 'profile') else review.reviewer.username
                        except:
                            reviewer_name = review.reviewer.username
                    ws.cell(row=row_num, column=9, value=reviewer_name).border = border
                    ws.cell(row=row_num, column=10, value=review.reviewed_at.strftime('%Y-%m-%d %H:%M') if review.reviewed_at else '').border = border
                    ws.cell(row=row_num, column=11, value=review.get_status_display()).border = border
                    ws.cell(row=row_num, column=12, value=review.reviewer_comment or '').border = border
                    row_num += 1
            else:
                ws.cell(row=row_num, column=1, value=item.id).border = border
                ws.cell(row=row_num, column=2, value=item.club.name).border = border
                ws.cell(row=row_num, column=3, value=item.submission_date.strftime('%Y-%m-%d')).border = border
                ws.cell(row=row_num, column=4, value=float(item.reimbursement_amount)).border = border
                ws.cell(row=row_num, column=5, value=item.get_status_display()).border = border
                ws.cell(row=row_num, column=6, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                ws.cell(row=row_num, column=7, value=item.resubmission_attempt).border = border
                for col in range(8, 13):
                    ws.cell(row=row_num, column=col, value='').border = border
                row_num += 1
        
        filename = f"报销记录-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
    elif tab_internal == 'activity_application':
        ws.title = "活动申请"
        headers = ['申请ID', '社团名称', '活动名称', '活动日期', '活动地点', '状态', '提交时间', '审核ID', '审核人', '审核时间', '审核状态', '审核意见']
        
        # 获取数据并应用筛选条件
        items = ActivityApplication.objects.all()
        if club_name:
            items = items.filter(club__name__icontains=club_name)
        if start_date:
            items = items.filter(submitted_at__gte=start_date)
        if end_date:
            items = items.filter(submitted_at__lte=end_date + ' 23:59:59')
        if status:
            items = items.filter(status=status)
        items = items.order_by('-submitted_at')
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 写入数据 - 包含所有历史审核记录
        from .models import ActivityApplicationHistory
        row_num = 2
        for item in items:
            reviews = ActivityApplicationHistory.objects.filter(activity_application=item).order_by('attempt_number')
            
            if reviews.exists():
                for review in reviews:
                    ws.cell(row=row_num, column=1, value=item.id).border = border
                    ws.cell(row=row_num, column=2, value=item.club.name).border = border
                    ws.cell(row=row_num, column=3, value=item.activity_name).border = border
                    ws.cell(row=row_num, column=4, value=item.activity_date.strftime('%Y-%m-%d')).border = border
                    ws.cell(row=row_num, column=5, value=item.activity_location).border = border
                    ws.cell(row=row_num, column=6, value=item.get_status_display()).border = border
                    ws.cell(row=row_num, column=7, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                    ws.cell(row=row_num, column=8, value=review.id).border = border
                    reviewer_name = ''
                    if review.reviewer:
                        try:
                            reviewer_name = review.reviewer.profile.get_full_name() if hasattr(review.reviewer, 'profile') else review.reviewer.username
                        except:
                            reviewer_name = review.reviewer.username
                    ws.cell(row=row_num, column=9, value=reviewer_name).border = border
                    ws.cell(row=row_num, column=10, value=review.reviewed_at.strftime('%Y-%m-%d %H:%M') if review.reviewed_at else '').border = border
                    ws.cell(row=row_num, column=11, value=review.get_status_display()).border = border
                    ws.cell(row=row_num, column=12, value=review.comment or '').border = border
                    row_num += 1
            else:
                ws.cell(row=row_num, column=1, value=item.id).border = border
                ws.cell(row=row_num, column=2, value=item.club.name).border = border
                ws.cell(row=row_num, column=3, value=item.activity_name).border = border
                ws.cell(row=row_num, column=4, value=item.activity_date.strftime('%Y-%m-%d')).border = border
                ws.cell(row=row_num, column=5, value=item.activity_location).border = border
                ws.cell(row=row_num, column=6, value=item.get_status_display()).border = border
                ws.cell(row=row_num, column=7, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
                for col in range(8, 13):
                    ws.cell(row=row_num, column=col, value='').border = border
                row_num += 1
        
        filename = f"活动申请-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
    elif tab_internal == 'president_transition':
        ws.title = "社长换届"
        headers = ['申请ID', '社团名称', '原社长', '新社长', '状态', '提交时间', '审核人', '审核时间', '审核意见']
        
        # 获取数据并应用筛选条件
        items = PresidentTransition.objects.all()
        if club_name:
            items = items.filter(club__name__icontains=club_name)
        if start_date:
            items = items.filter(submitted_at__gte=start_date)
        if end_date:
            items = items.filter(submitted_at__lte=end_date + ' 23:59:59')
        if status:
            items = items.filter(status=status)
        items = items.order_by('-submitted_at')
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 写入数据
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item.id).border = border
            ws.cell(row=row_num, column=2, value=item.club.name).border = border
            try:
                old_president_name = item.old_president.profile.get_full_name() if hasattr(item.old_president, 'profile') else item.old_president.username
            except:
                old_president_name = item.old_president.username if item.old_president else ''
            ws.cell(row=row_num, column=3, value=old_president_name).border = border
            try:
                new_president_name = item.new_president.profile.get_full_name() if hasattr(item.new_president, 'profile') else item.new_president.username
            except:
                new_president_name = item.new_president.username if item.new_president else ''
            ws.cell(row=row_num, column=4, value=new_president_name).border = border
            ws.cell(row=row_num, column=5, value=item.get_status_display()).border = border
            ws.cell(row=row_num, column=6, value=item.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
            reviewer_name = ''
            if item.reviewer:
                try:
                    reviewer_name = item.reviewer.profile.get_full_name() if hasattr(item.reviewer, 'profile') else item.reviewer.username
                except:
                    reviewer_name = item.reviewer.username
            ws.cell(row=row_num, column=7, value=reviewer_name).border = border
            ws.cell(row=row_num, column=8, value=item.reviewed_at.strftime('%Y-%m-%d %H:%M') if item.reviewed_at else '').border = border
            ws.cell(row=row_num, column=9, value=item.reviewer_comment or '').border = border
        
        filename = f"社长换届-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    
    else:
        messages.error(request, '无效的数据类型，无法导出')
        return redirect('clubs:staff_audit_center', tab=tab)
    
    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 18
    
    # 生成响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(filename)}'
    
    wb.save(response)
    return response
