import base64
import json
import zipfile
from datetime import date, time
from hashlib import md5
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
from urllib.parse import urlparse

from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.files.storage import default_storage
from django.db import IntegrityError, connection, transaction
from django.http import QueryDict
from django.test import TestCase, override_settings
from django.test.utils import CaptureQueriesContext
from django.urls import URLPattern, URLResolver, get_resolver, reverse
from django.utils import timezone
from django.utils.datastructures import MultiValueDict

from ..avatar_utils import clear_avatar_settings_cache, get_profile_avatar_url
from ..models import (
    Announcement,
    ChannelExampleFile,
    Club,
    ClubMember,
    EmailVerificationCode,
    FormChannel,
    FormCycle,
    FormField,
    FormFieldValue,
    FormSubmission,
    FormSubmissionReview,
    Officer,
    RegistrationToken,
    Room,
    RoomBooking,
    SiteSettings,
    SMTPConfig,
    StaffClubRelation,
    FileBlob,
    TimeSlot,
    UserProfile,
)
from ..services.booking_service import BookingConflictError, create_room_booking
from ..services.registration_service import (
    RegistrationTokenUnavailable,
    register_member_with_token,
)
from ..storage_backends import bind_client_md5_from_post
from ..upload_security import validate_upload


class AvatarServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username='avatar-admin',
            email='',
            password='test-password',
        )
        self.config = SiteSettings.get_settings()
        self.config.cravatar_enabled = True
        self.config.save(update_fields=['cravatar_enabled'])
        clear_avatar_settings_cache()
        self.client.force_login(self.user)

    def tearDown(self):
        clear_avatar_settings_cache()

    def test_cravatar_is_an_individual_user_choice(self):
        profile = self.user.profile
        profile.avatar_email = 'Person@Example.COM'
        profile.avatar_source = 'cravatar'
        profile.save(update_fields=['avatar_email', 'avatar_source'])
        expected_hash = md5(b'person@example.com').hexdigest()

        self.assertIn(expected_hash, get_profile_avatar_url(profile))

    def test_cravatar_proxy_rejects_invalid_digest(self):
        response = self.client.get('/cravatar/not-a-digest/', secure=True)
        self.assertEqual(response.status_code, 404)

    def test_cravatar_proxy_serves_cached_with_immutable_header(self):
        digest = 'a' * 32
        cache.set(f'cravatar:{digest}:160:mp:g', {'content': b'png-bytes', 'content_type': 'image/png'})

        response = self.client.get(f'/cravatar/{digest}/', secure=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Cache-Control'], 'public, max-age=31536000, immutable')
        self.assertEqual(response.content, b'png-bytes')

    def test_is_image_file_supports_model_file_objects(self):
        from django.core.files.base import ContentFile
        from ..templatetags.common_tags import is_image_file
        from ..models import FormUploadedFile

        channel = FormChannel.objects.create(name='图片通道', slug='image-channel', is_active=True)
        submission = FormSubmission.objects.create(
            channel=channel,
            club=Club.objects.create(name='图片社团', founded_date=date(2026, 1, 1)),
            submitter=self.user,
        )
        field = FormField.objects.create(channel=channel, field_key='img', label='图片', field_type='file')
        uploaded = FormUploadedFile.objects.create(
            submission=submission,
            field=field,
            file=ContentFile(b'png', name='photo.png'),
            original_name='photo.png',
        )

        self.assertTrue(is_image_file(uploaded))

    def test_disabled_cravatar_does_not_override_local_avatar(self):
        profile = self.user.profile
        profile.avatar_email = 'person@example.com'
        profile.avatar_source = 'cravatar'
        profile.save(update_fields=['avatar_email', 'avatar_source'])
        self.config.cravatar_enabled = False
        self.config.save(update_fields=['cravatar_enabled'])
        clear_avatar_settings_cache()

        self.assertEqual(get_profile_avatar_url(profile), '')

    def test_admin_only_controls_cravatar_availability(self):
        response = self.client.post(
            reverse('clubs:site_settings'),
            {'form_type': 'avatar_settings', 'cravatar_enabled': 'on'},
            secure=True,
        )

        self.assertRedirects(response, reverse('clubs:site_settings'), fetch_redirect_response=False)
        self.config.refresh_from_db()
        self.assertTrue(self.config.cravatar_enabled)

    def test_avatar_upload_returns_json_when_accept_header_requests_it(self):
        response = self.client.post(
            reverse('clubs:edit_profile'),
            {'action': 'upload_avatar'},
            HTTP_ACCEPT='application/json',
            secure=True,
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['message'], '请选择图片文件')

    def test_upload_avatar_deletes_old_file(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from PIL import Image

        from ..storage_backends import ClubStorage

        def upload_png():
            buffer = BytesIO()
            Image.new('RGB', (32, 32), 'red').save(buffer, format='PNG')
            response = self.client.post(
                reverse('clubs:edit_profile'),
                {
                    'action': 'upload_avatar',
                    'avatar': SimpleUploadedFile('avatar.png', buffer.getvalue(), content_type='image/png'),
                },
                HTTP_ACCEPT='application/json',
                HTTP_X_REQUESTED_WITH='XMLHttpRequest',
                secure=True,
            )
            return response

        with (
            patch.object(
                ClubStorage,
                'save',
                side_effect=lambda name, content, max_length=None: name,
            ) as save_mock,
            patch.object(ClubStorage, 'delete') as delete_mock,
        ):
            first_response = upload_png()
            self.assertEqual(first_response.status_code, 200)
            self.user.profile.refresh_from_db()
            first_name = self.user.profile.avatar.name

            second_response = upload_png()
            self.assertEqual(second_response.status_code, 200)
            self.user.profile.refresh_from_db()
            second_name = self.user.profile.avatar.name

        self.assertNotEqual(first_name, second_name)
        self.assertEqual(save_mock.call_count, 2)
        delete_mock.assert_called_once_with(first_name)

    def test_department_members_visible_to_logged_in_users(self):
        from ..models import Department

        department = Department.objects.create(name='测试部门')
        member = User.objects.create_user(username='plain-member', password='test-password')
        UserProfile.objects.create(user=member, role='member', status='approved')
        self.client.force_login(member)

        response = self.client.get(
            reverse('clubs:get_department_members', args=[department.pk]),
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['name'], '测试部门')

    def test_department_members_requires_login(self):
        from ..models import Department

        department = Department.objects.create(name='测试部门')
        self.client.logout()

        response = self.client.get(
            reverse('clubs:get_department_members', args=[department.pk]),
            secure=True,
        )

        self.assertEqual(response.status_code, 302)

    def test_admin_room_add_rejects_invalid_capacity(self):
        staff = User.objects.create_user(username='audit-staff', password='test-password')
        UserProfile.objects.create(user=staff, role='staff', status='approved')
        self.client.force_login(staff)

        response = self.client.post(
            reverse('clubs:admin_room_add'),
            {
                'name': '测试房间',
                'capacity': 'not-a-number',
                'status': 'available',
            },
            secure=True,
        )

        self.assertRedirects(response, reverse('clubs:admin_room_add'), fetch_redirect_response=False)
        self.assertEqual(Room.objects.count(), 0)

    @override_settings(USE_X_FORWARDED_FOR=True)
    def test_login_lockout_uses_forwarded_for_ip(self):
        from ..views import auth as auth_views

        self.client.logout()
        login_url = reverse('clubs:login')
        fake_ip = '203.0.113.99'
        for _ in range(auth_views.MAX_LOGIN_ATTEMPTS):
            self.client.post(
                login_url,
                {'username': 'unknown-user-xyz', 'password': 'wrong-password'},
                HTTP_X_FORWARDED_FOR=fake_ip,
                secure=True,
            )

        response = self.client.post(
            login_url,
            {'username': 'unknown-user-xyz', 'password': 'wrong-password'},
            HTTP_X_FORWARDED_FOR=fake_ip,
            secure=True,
        )

        self.assertContains(response, '登录尝试过多')

    @override_settings(USE_X_FORWARDED_FOR=True)
    def test_login_lockout_uses_leftmost_forwarded_ip(self):
        from ..views import auth as auth_views

        self.client.logout()
        login_url = reverse('clubs:login')
        # 多级代理时 XFF 形如“客户端IP, 代理IP”，限速必须按最左侧客户端 IP 计数。
        for index in range(auth_views.MAX_LOGIN_ATTEMPTS):
            self.client.post(
                login_url,
                {'username': f'unknown-user-multi-{index}', 'password': 'wrong-password'},
                HTTP_X_FORWARDED_FOR=f'10.0.0.{index}, 198.51.100.9',
                secure=True,
            )

        response = self.client.post(
            login_url,
            {'username': 'unknown-user-multi-final', 'password': 'wrong-password'},
            HTTP_X_FORWARDED_FOR='10.0.0.99, 198.51.100.9',
            secure=True,
        )

        # 每次来源 IP 不同，不应触发 IP 限速。
        self.assertNotContains(response, '登录尝试过多')

    def test_login_lockout_ignores_forwarded_for_without_proxy(self):
        from ..views import auth as auth_views

        self.client.logout()
        login_url = reverse('clubs:login')
        for index in range(auth_views.MAX_LOGIN_ATTEMPTS):
            self.client.post(
                login_url,
                {'username': 'unknown-user-abc', 'password': 'wrong-password'},
                HTTP_X_FORWARDED_FOR=f'10.0.0.{index}',
                secure=True,
            )

        response = self.client.post(
            login_url,
            {'username': 'unknown-user-abc', 'password': 'wrong-password'},
            HTTP_X_FORWARDED_FOR='10.0.0.99',
            secure=True,
        )

        self.assertContains(response, '登录尝试过多')

    def test_build_external_url_respects_proxy_switch(self):
        from django.test import RequestFactory

        from ..views.core import _build_external_url

        request = RequestFactory().get('/some/path')
        request.META['HTTP_X_FORWARDED_PROTO'] = 'https'
        request.META['HTTP_X_FORWARDED_HOST'] = 'public.example.com'

        with override_settings(USE_X_FORWARDED_PROTO=False, USE_X_FORWARDED_HOST=False):
            url = _build_external_url(request, '/path')
            self.assertTrue(url.startswith('http://testserver'))

        with override_settings(USE_X_FORWARDED_PROTO=True, USE_X_FORWARDED_HOST=True):
            url = _build_external_url(request, '/path')
            parsed = urlparse(url)
            self.assertEqual(parsed.scheme, 'https')
            self.assertEqual(parsed.hostname, 'public.example.com')

    def test_office_preview_url_rejects_non_http_schemes(self):
        from django.core.files.base import ContentFile
        from django.test import RequestFactory

        from ..views.core import _office_preview_url
        from ..models import FormField, FormSubmission, FormUploadedFile

        channel = FormChannel.objects.create(
            name='预览测试通道',
            slug='preview-test',
            is_active=True,
            publish_status='published',
        )
        submission = FormSubmission.objects.create(
            channel=channel,
            club=Club.objects.create(name='预览测试社团', founded_date=date(2026, 1, 1)),
            submitter=self.user,
        )
        field = FormField.objects.create(channel=channel, field_key='f', label='附件', field_type='file')
        uploaded = FormUploadedFile.objects.create(
            submission=submission,
            field=field,
            file=ContentFile(b'docx-bytes', name='doc.docx'),
            original_name='doc.docx',
        )
        uploaded.file.name = 'file:///etc/passwd.docx'

        request = RequestFactory().get('/')
        self.assertEqual(_office_preview_url(request, uploaded), '')

    def test_storage_name_sanitization(self):
        from ..storage_backends import _sanitize_storage_name

        self.assertEqual(_sanitize_storage_name('file:///etc/passwd.docx'), 'etc/passwd.docx')
        self.assertEqual(_sanitize_storage_name('/abs/path/x.docx'), 'abs/path/x.docx')
        self.assertEqual(_sanitize_storage_name('form_submissions/abc/x.docx'), 'form_submissions/abc/x.docx')
        self.assertEqual(_sanitize_storage_name('a/../b.docx'), 'b.docx')
        self.assertEqual(_sanitize_storage_name(''), 'unnamed')

    def test_admin_can_change_third_party_cdn(self):
        response = self.client.post(
            reverse('clubs:site_settings'),
            {
                'form_type': 'cdn_settings',
                'third_party_cdn_base_url': 'https://mirrors.sustech.edu.cn/cdnjs/',
                'third_party_cdn_sri': json.dumps({
                    'chartjs': 'sha384-XcdcwHqIPULERb2yDEM4R0XaQKU3YnDsrTmjACBZyfdVVqjh6xQ4/DCMd7XLcA6Y',
                    'swiper_js': 'sha384-T6qkM4ANslBL/pKcwNUeB0bpsiI6pkXXzwrl7Avc6FXEC/UZaXAeBpZZ2zQ3Zbez',
                    'swiper_css': 'sha384-eKrJLy2KlZuvuza/yNmSyFUE2Qb5aehRlXikp6XUOxXVw5pOQBb5n1C0UOcCnAJb',
                    'cropper_js': 'sha384-jrOgQzBlDeUNdmQn3rUt/PZD+pdcRBdWd/HWRqRo+n2OR2QtGyjSaJC0GiCeH+ir',
                    'cropper_css': 'sha384-6LFfkTKLRlzFtgx8xsWyBdKGpcMMQTkv+dB7rAbugeJAu1Ym2q1Aji1cjHBG12Xh',
                }),
            },
            secure=True,
        )

        self.assertRedirects(response, reverse('clubs:site_settings'), fetch_redirect_response=False)
        self.config.refresh_from_db()
        self.assertEqual(
            self.config.third_party_cdn_base_url,
            'https://mirrors.sustech.edu.cn/cdnjs',
        )
        self.assertTrue(self.config.third_party_cdn_sri['chartjs'].startswith('sha384-'))

    def test_admin_can_change_site_name_and_homepage_title(self):
        response = self.client.post(
            reverse('clubs:site_settings'),
            {
                'form_type': 'site_info_settings',
                'site_name': '测试站点',
                'homepage_title': '测试首页标题',
                'homepage_subtitle': '测试首页副标题',
            },
            secure=True,
        )

        self.assertRedirects(response, reverse('clubs:site_settings'), fetch_redirect_response=False)
        self.config.refresh_from_db()
        self.assertEqual(self.config.site_name, '测试站点')
        self.assertEqual(self.config.homepage_title, '测试首页标题')
        self.assertEqual(self.config.homepage_subtitle, '测试首页副标题')

    def test_third_party_cdn_rejects_urls_with_query_parameters(self):
        original_url = self.config.third_party_cdn_base_url
        response = self.client.post(
            reverse('clubs:site_settings'),
            {
                'form_type': 'cdn_settings',
                'third_party_cdn_base_url': 'https://cdn.example.com?redirect=elsewhere',
            },
            secure=True,
        )

        self.assertRedirects(response, reverse('clubs:site_settings'), fetch_redirect_response=False)
        self.config.refresh_from_db()
        self.assertEqual(self.config.third_party_cdn_base_url, original_url)

    @patch('clubs.avatar_utils.cravatar_exists', return_value=True)
    def test_existing_cravatar_is_saved_and_selected(self, exists_mock):
        response = self.client.post(
            reverse('clubs:edit_profile'),
            {'action': 'use_cravatar', 'avatar_email': 'Avatar@Example.com'},
            secure=True,
        )

        self.assertRedirects(
            response,
            f"{reverse('clubs:edit_profile')}?tab=avatar",
            fetch_redirect_response=False,
        )
        self.user.profile.refresh_from_db()
        self.assertEqual(self.user.profile.avatar_email, 'avatar@example.com')
        self.assertEqual(self.user.profile.avatar_source, 'cravatar')
        exists_mock.assert_called_once_with('avatar@example.com')

    @patch('clubs.avatar_utils.cravatar_exists', return_value=False)
    def test_missing_cravatar_keeps_current_selection(self, exists_mock):
        response = self.client.post(
            reverse('clubs:edit_profile'),
            {'action': 'use_cravatar', 'avatar_email': 'missing@example.com'},
            secure=True,
        )

        self.assertRedirects(
            response,
            f"{reverse('clubs:edit_profile')}?tab=avatar",
            fetch_redirect_response=False,
        )
        self.user.profile.refresh_from_db()
        self.assertEqual(self.user.profile.avatar_email, '')
        self.assertEqual(self.user.profile.avatar_source, 'local')
        exists_mock.assert_called_once_with('missing@example.com')

    def test_user_can_switch_back_to_local(self):
        profile = self.user.profile
        profile.avatar_source = 'cravatar'
        profile.avatar_email = 'person@example.com'
        profile.save(update_fields=['avatar_source', 'avatar_email'])

        response = self.client.post(
            reverse('clubs:edit_profile'),
            {'action': 'use_local_avatar'},
            secure=True,
        )

        self.assertRedirects(
            response,
            f"{reverse('clubs:edit_profile')}?tab=avatar",
            fetch_redirect_response=False,
        )
        profile.refresh_from_db()
        self.assertEqual(profile.avatar_source, 'local')

    def test_local_avatar_response_has_immutable_browser_cache(self):
        with TemporaryDirectory() as media_root:
            avatar_path = Path(media_root, 'avatars', '2026', '08', 'avatar.jpg')
            avatar_path.parent.mkdir(parents=True)
            avatar_path.write_bytes(b'avatar-image')

            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get('/media/avatars/2026/08/avatar.jpg', secure=True)

            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                response.headers['Cache-Control'],
                'public, max-age=31536000, immutable',
            )


class StaffRegistrationReviewTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username='review-admin',
            email='',
            password='test-password',
        )
        UserProfile.objects.get_or_create(
            user=self.admin,
            defaults={'role': 'admin', 'status': 'approved'},
        )
        self.client.force_login(self.admin)

    @staticmethod
    def create_pending_staff(username):
        user = User.objects.create_user(username=username, password='test-password')
        UserProfile.objects.create(user=user, role='staff', status='pending')
        return user

    def test_review_accepts_current_decision_field(self):
        staff = self.create_pending_staff('pending-current')

        response = self.client.post(
            reverse('clubs:review_staff_registration', args=[staff.pk]),
            {'decision': 'approved'},
            secure=True,
        )

        self.assertRedirects(response, reverse('clubs:manage_users'), fetch_redirect_response=False)
        staff.profile.refresh_from_db()
        self.assertEqual(staff.profile.status, 'approved')

    def test_review_rejects_already_processed_submission(self):
        first_reviewer = self.create_pending_staff('reviewer-first')
        first_reviewer.profile.status = 'approved'
        first_reviewer.profile.save(update_fields=['status'])
        channel = FormChannel.objects.create(
            name='审核通道',
            slug='review-channel',
            is_active=True,
            publish_status='published',
        )
        club = Club.objects.create(name='审核社团', founded_date=date(2026, 1, 1))
        submission = FormSubmission.objects.create(
            channel=channel,
            club=club,
            submitter=first_reviewer,
            status='approved',
        )
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=first_reviewer,
            status='approved',
            submission_attempt=1,
        )

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'decision': 'approved', 'comment': '重复审核'},
            secure=True,
        )

        self.assertRedirects(
            response,
            reverse('clubs:staff_audit_center', args=[submission.channel.slug]),
            fetch_redirect_response=False,
        )
        followed = self.client.get(response.url, secure=True)
        self.assertContains(followed, '无需重复审核')
        submission.refresh_from_db()
        self.assertEqual(submission.status, 'approved')

    def test_user_management_shows_pending_review_count(self):
        self.create_pending_staff('pending-count')
        approved = self.create_pending_staff('approved-count')
        approved.profile.status = 'approved'
        approved.profile.save(update_fields=['status', 'updated_at'])

        response = self.client.get(reverse('clubs:manage_users'), secure=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['pending_review_users'], 1)
        self.assertContains(response, '1 个待审核')
        self.assertContains(response, '?role=staff&amp;status=pending')

    def test_batch_delete_users(self):
        first = User.objects.create_user(username='batch-del-1', password='test-password')
        UserProfile.objects.create(user=first, role='member', status='approved')
        second = User.objects.create_user(username='batch-del-2', password='test-password')
        UserProfile.objects.create(user=second, role='member', status='approved')

        response = self.client.post(
            reverse('clubs:manage_users'),
            {'action': 'batch_delete', 'user_ids': f'{first.pk},{second.pk}'},
            secure=True,
        )

        self.assertRedirects(response, reverse('clubs:manage_users'), fetch_redirect_response=False)
        self.assertFalse(User.objects.filter(pk__in=[first.pk, second.pk]).exists())

    def test_batch_toggle_active_users(self):
        first = User.objects.create_user(username='batch-tog-1', password='test-password')
        UserProfile.objects.create(user=first, role='member', status='approved')
        second = User.objects.create_user(username='batch-tog-2', password='test-password')
        UserProfile.objects.create(user=second, role='member', status='approved')

        self.client.post(
            reverse('clubs:manage_users'),
            {'action': 'batch_disable', 'user_ids': f'{first.pk},{second.pk}'},
            secure=True,
        )
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertFalse(first.is_active)
        self.assertFalse(second.is_active)

        self.client.post(
            reverse('clubs:manage_users'),
            {'action': 'batch_enable', 'user_ids': f'{first.pk},{second.pk}'},
            secure=True,
        )
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertTrue(first.is_active)
        self.assertTrue(second.is_active)

    def test_batch_delete_skips_superuser_and_self(self):
        target = User.objects.create_user(username='batch-skip-1', password='test-password')
        UserProfile.objects.create(user=target, role='member', status='approved')

        self.client.post(
            reverse('clubs:manage_users'),
            {'action': 'batch_delete', 'user_ids': f'{target.pk},{self.admin.pk}'},
            secure=True,
        )

        self.assertFalse(User.objects.filter(pk=target.pk).exists())
        self.assertTrue(User.objects.filter(pk=self.admin.pk).exists())

    def test_manage_announcements_page_and_toggle(self):
        announcement = Announcement.objects.create(
            title='测试公告',
            content='公告内容',
            status='draft',
            created_by=self.admin,
        )

        response = self.client.get(reverse('clubs:manage_announcements'), secure=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '测试公告')

        self.client.post(
            reverse('clubs:toggle_announcement_status', args=[announcement.pk]),
            secure=True,
        )
        announcement.refresh_from_db()
        self.assertEqual(announcement.status, 'published')

    def test_time_slot_import_template_and_import(self):
        template_response = self.client.get(reverse('clubs:download_time_slot_import_template'), secure=True)
        self.assertEqual(template_response.status_code, 200)
        self.assertContains(template_response, '显示名称')

        csv_content = '显示名称,开始时间,结束时间,状态\n第1-2节,08:15,09:55,启用\n午休,11:40,13:00,启用\n'
        upload = SimpleUploadedFile(
            'time_slots.csv',
            csv_content.encode('utf-8-sig'),
            content_type='text/csv',
        )
        response = self.client.post(
            reverse('clubs:import_time_slots_csv'),
            {'csv_file': upload},
            HTTP_ACCEPT='application/json',
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertEqual(TimeSlot.objects.filter(label='第1-2节').count(), 1)
        self.assertEqual(TimeSlot.objects.filter(label='午休').count(), 1)

    def test_weekly_export_uses_monday_and_configured_slots(self):
        from openpyxl import load_workbook

        room = Room.objects.create(name='导出测试房间', capacity=10)
        TimeSlot.objects.create(label='自定义时段', start_time=time(9, 0), end_time=time(10, 0), is_active=True)

        response = self.client.get(
            reverse('clubs:export_room_bookings_weekly'),
            {'room_id': room.pk, 'week_start': '2026-08-05'},
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertIn('2026年08月03日', worksheet['A1'].value)
        labels = [worksheet.cell(row=row, column=1).value for row in range(3, worksheet.max_row + 1)]
        self.assertTrue(any(label and '自定义时段' in label for label in labels))
        self.assertTrue(any(label and '09:00-10:00' in label for label in labels))

    def test_save_form_channel_accepts_multiple_example_files(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        channel = FormChannel.objects.create(
            name='示例通道',
            slug='example-channel',
            is_active=True,
            publish_status='draft',
        )

        def make_docx(name):
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as archive:
                archive.writestr('[Content_Types].xml', '<Types/>')
            return SimpleUploadedFile(
                name,
                zip_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            )

        response = self.client.post(
            reverse('clubs:edit_form_channel', args=[channel.pk]),
            {
                'name': '示例通道',
                'slug': 'example-channel',
                'icon': 'description',
                'order': '0',
                'required_approval_count': '1',
                'publish_status': 'draft',
                'example_files': [make_docx('a.docx'), make_docx('b.docx')],
            },
            secure=True,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(ChannelExampleFile.objects.filter(channel=channel).count(), 2)


class BookingServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser('booking-admin', '', 'password')
        self.room = Room.objects.create(name='事务测试房间', capacity=20)

    def create_booking(self, start=time(9), end=time(10)):
        return create_room_booking(
            room_id=self.room.pk,
            user=self.user,
            club=None,
            booking_date=date(2026, 8, 10),
            start_time=start,
            end_time=end,
            purpose='测试预约',
            participant_count=5,
            contact_phone='13800000000',
        )

    def test_overlapping_booking_is_rejected(self):
        self.create_booking()

        with self.assertRaises(BookingConflictError):
            self.create_booking(time(9, 30), time(10, 30))

        self.assertEqual(RoomBooking.objects.count(), 1)

    def test_adjacent_booking_is_allowed(self):
        self.create_booking()
        self.create_booking(time(10), time(11))

        self.assertEqual(RoomBooking.objects.count(), 2)

    def test_database_rejects_non_positive_participants(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            RoomBooking.objects.create(
                room=self.room,
                user=self.user,
                booking_date=date(2026, 8, 10),
                start_time=time(9),
                end_time=time(10),
                purpose='无效预约',
                participant_count=0,
                contact_phone='13800000000',
            )

    def test_booking_cancellation_requires_confirmed_post(self):
        booking = self.create_booking()
        self.client.force_login(self.user)

        confirmation = self.client.get(
            reverse('clubs:delete_room_booking', args=[booking.pk]),
            secure=True,
        )
        self.assertEqual(confirmation.status_code, 200)
        self.assertTrue(RoomBooking.objects.filter(pk=booking.pk).exists())

        response = self.client.post(
            reverse('clubs:delete_room_booking', args=[booking.pk]),
            secure=True,
        )
        self.assertRedirects(
            response,
            reverse('clubs:my_room_bookings'),
            fetch_redirect_response=False,
        )
        self.assertFalse(RoomBooking.objects.filter(pk=booking.pk).exists())


class RegistrationTokenTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser('token-admin', '', 'password')
        self.club = Club.objects.create(
            name='令牌测试社团',
            founded_date=date(2020, 1, 1),
        )
        self.token = RegistrationToken.create_for_club(
            club=self.club,
            created_by=self.user,
            minutes=10,
            max_uses=1,
        )

    def test_last_token_use_cannot_be_consumed_twice(self):
        self.token.mark_used()

        with self.assertRaises(ValidationError):
            self.token.mark_used()

        self.token.refresh_from_db()
        self.assertEqual(self.token.used_count, 1)

    def test_generated_tokens_are_unique_and_url_safe(self):
        first = RegistrationToken.generate_code()
        second = RegistrationToken.generate_code()

        self.assertNotEqual(first, second)
        self.assertGreaterEqual(len(first), 32)
        self.assertRegex(first, r'^[A-Za-z0-9_-]+$')

    def test_registration_and_last_use_are_one_transaction(self):
        registered_user, _ = register_member_with_token(
            token_id=self.token.pk,
            user=None,
            existing_account=False,
            username='new-member',
            password='member-password',
            email='member@example.com',
            profile_data={
                'real_name': '新社员',
                'student_id': '20260001',
                'gender': 'other',
                'college': '测试学院',
                'class_name': '测试班',
                'phone': '13800000000',
                'qq': '',
                'wechat': 'new-member',
            },
        )

        self.assertTrue(ClubMember.objects.filter(
            club=self.club,
            user_profile=registered_user.profile,
        ).exists())
        self.token.refresh_from_db()
        self.assertEqual(self.token.used_count, 1)

        with self.assertRaises(RegistrationTokenUnavailable):
            register_member_with_token(
                token_id=self.token.pk,
                user=None,
                existing_account=False,
                username='should-not-exist',
                password='member-password',
                email='unused@example.com',
                profile_data={
                    'real_name': '不会创建',
                    'student_id': '20260002',
                    'gender': 'other',
                    'college': '测试学院',
                    'class_name': '测试班',
                    'phone': '13800000001',
                    'qq': '',
                    'wechat': 'unused',
                },
            )

        self.assertFalse(User.objects.filter(username='should-not-exist').exists())


class SecurityInfrastructureTests(TestCase):
    def test_service_credentials_are_encrypted_at_rest(self):
        config = SMTPConfig.objects.create(
            provider='custom',
            smtp_host='smtp.example.com',
            smtp_port=587,
            sender_email='sender@example.com',
            sender_password='plain-secret',
        )
        with connection.cursor() as cursor:
            cursor.execute('SELECT sender_password FROM clubs_smtpconfig WHERE id = %s', [config.pk])
            stored = cursor.fetchone()[0]

        self.assertNotEqual(stored, 'plain-secret')
        self.assertTrue(stored.startswith('enc:v1:'))
        config.refresh_from_db()
        self.assertEqual(config.sender_password, 'plain-secret')

    def test_upload_signature_must_match_extension(self):
        upload = SimpleUploadedFile('fake.png', b'not-a-png', content_type='image/png')

        error = validate_upload(upload, field_name='图片', allowed_extensions={'.png'})

        self.assertIn('实际内容', error)

    def test_zip_path_traversal_is_rejected(self):
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, 'w') as archive:
            archive.writestr('../escape.txt', 'unsafe')
        upload = SimpleUploadedFile('unsafe.zip', buffer.getvalue(), content_type='application/zip')

        error = validate_upload(upload, field_name='压缩包', allowed_extensions={'.zip'})

        self.assertIn('不安全的路径', error)

class AccountSecurityAndHistoryTests(TestCase):
    def setUp(self):
        cache.clear()
        admin = User.objects.create_superuser(
            username='security-admin',
            email='',
            password='test-password',
        )
        UserProfile.objects.get_or_create(
            user=admin,
            defaults={'role': 'admin', 'status': 'approved'},
        )

    def test_staff_cannot_self_assign_responsible_club(self):
        club = Club.objects.create(name='受限社团', founded_date=date(2026, 1, 1))
        staff = User.objects.create_user('staff-self-assign', password='test-password')
        UserProfile.objects.create(user=staff, role='staff', status='approved')
        self.client.force_login(staff)

        response = self.client.post(
            reverse('clubs:manage_staff_clubs'),
            {'club_ids': [club.pk]},
            secure=True,
        )

        self.assertRedirects(
            response,
            reverse('clubs:manage_staff_clubs'),
            fetch_redirect_response=False,
        )
        self.assertFalse(
            StaffClubRelation.objects.filter(
                staff=staff.profile,
                club=club,
                is_active=True,
            ).exists()
        )

    def test_email_verification_locks_after_five_wrong_attempts(self):
        user = User.objects.create_user('verify-lock', password='test-password')
        UserProfile.objects.create(user=user, role='member', status='approved')
        code = EmailVerificationCode.generate_code()
        EmailVerificationCode.objects.create(
            user=user,
            email='verify-lock@example.com',
            code=code,
            expires_at=timezone.now() + timezone.timedelta(minutes=15),
        )
        self.client.force_login(user)

        for _ in range(5):
            self.client.post(
                reverse('clubs:verify_email'),
                {'code': '000000'},
                secure=True,
            )

        verification = EmailVerificationCode.objects.get(user=user)
        self.assertGreaterEqual(verification.failed_attempts, 5)
        success, message = verification.verify(code)
        self.assertFalse(success)
        self.assertIn('错误次数过多', message)

    @patch('clubs.email_utils.send_verification_email', return_value=(True, '已发送'))
    def test_resend_verification_code_has_cooldown(self, mock_send):
        user = User.objects.create_user('resend-limit', password='test-password')
        UserProfile.objects.create(user=user, role='member', status='approved')
        EmailVerificationCode.objects.create(
            user=user,
            email='resend-limit@example.com',
            code=EmailVerificationCode.generate_code(),
            expires_at=timezone.now() + timezone.timedelta(minutes=15),
        )
        self.client.force_login(user)

        self.client.post(reverse('clubs:resend_verification_code'), secure=True)
        self.client.post(reverse('clubs:resend_verification_code'), secure=True)

        self.assertEqual(mock_send.call_count, 1)

    def test_deleting_reviewer_keeps_review_history(self):
        channel = FormChannel.objects.create(
            name='历史保留通道',
            slug='history-review-channel',
            is_active=True,
            publish_status='published',
        )
        club = Club.objects.create(name='历史保留社团', founded_date=date(2026, 1, 1))
        reviewer = User.objects.create_user('deleted-reviewer', password='test-password')
        UserProfile.objects.create(user=reviewer, role='staff', status='approved')
        submitter = User.objects.create_user('history-submitter', password='test-password')
        UserProfile.objects.create(user=submitter, role='member', status='approved')
        submission = FormSubmission.objects.create(
            channel=channel,
            club=club,
            submitter=submitter,
            status='approved',
        )
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=reviewer,
            status='approved',
            submission_attempt=1,
        )

        reviewer.delete()

        review = FormSubmissionReview.objects.get(submission=submission)
        self.assertIsNone(review.reviewer)
        self.assertEqual(submission.reviewer_count, 1)
        self.assertEqual(submission.approved_review_count(), 1)


class StaticPageSmokeTests(TestCase):
    """Render every named, parameter-free page and fail on server errors."""

    def setUp(self):
        self.admin = User.objects.create_superuser(
            username='page-smoke-admin',
            email='admin@example.com',
            password='test-password',
        )

    def _named_static_patterns(self, patterns=None, namespace=''):
        patterns = patterns or get_resolver().url_patterns
        for pattern in patterns:
            if isinstance(pattern, URLResolver):
                child_namespace = namespace
                if pattern.namespace:
                    child_namespace = f'{namespace}:{pattern.namespace}' if namespace else pattern.namespace
                yield from self._named_static_patterns(pattern.url_patterns, child_namespace)
            elif isinstance(pattern, URLPattern) and pattern.name and '<' not in str(pattern.pattern):
                name = f'{namespace}:{pattern.name}' if namespace else pattern.name
                yield name

    def test_named_static_pages_do_not_raise_server_errors(self):
        excluded = {'admin:index', 'clubs:oobe_setup'}
        for name in sorted(set(self._named_static_patterns())):
            if name in excluded:
                continue
            with self.subTest(route=name):
                self.client.force_login(self.admin)
                response = self.client.get(reverse(name), follow=False, secure=True)
                self.assertLess(response.status_code, 500, f'{name} returned {response.status_code}')

    def test_form_channel_admin_exposes_status_overview_and_search(self):
        initial_count = FormChannel.objects.count()
        FormChannel.objects.create(
            name='界面测试通道',
            slug='ui-test-channel',
            publish_status='published',
            is_active=True,
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse('clubs:manage_form_channels'), secure=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['channel_summary']['total'], initial_count + 1)
        self.assertGreaterEqual(response.context['channel_summary']['published'], 1)
        self.assertContains(response, 'js-channel-search')
        self.assertContains(response, 'rail-search-control')
        self.assertContains(response, 'aria-label="搜索通道名称或标识"')
        self.assertContains(response, '通道状态概览')

    def test_new_form_channel_page_exposes_material_icon_preview_and_reference(self):
        self.client.force_login(self.admin)

        response = self.client.get(f"{reverse('clubs:manage_form_channels')}?new=1", secure=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'js-material-icon-preview')
        self.assertContains(response, 'js-material-icon-input')
        self.assertContains(response, '正在预览：description')
        self.assertContains(response, 'https://mui.com/material-ui/material-icons/')

    def test_staff_management_query_count_does_not_scale_per_club(self):
        profile = self.admin.profile
        profile.role = 'admin'
        profile.save(update_fields=['role'])
        for action, slug in [('annual_review', 'query-annual'), ('club_registration', 'query-registration')]:
            channel = FormChannel.objects.create(
                name=slug,
                slug=slug,
                builtin_action=action,
                cycle_type='count',
                submission_policy='once_per_cycle',
                allow_staff_toggle=True,
            )
            FormCycle.objects.create(channel=channel, name='性能测试周期', created_by=self.admin)

        def add_clubs(start, stop):
            for index in range(start, stop):
                president = User.objects.create(username=f'query-president-{index}')
                president_profile = UserProfile.objects.create(
                    user=president,
                    role='president',
                    real_name=f'性能社长 {index:02d}',
                )
                club = Club.objects.create(
                    name=f'查询性能社团 {index:02d}',
                    description='查询性能测试',
                    founded_date=date.today(),
                    members_count=1,
                )
                Officer.objects.create(
                    club=club,
                    user_profile=president_profile,
                    position='president',
                    appointed_date=date.today(),
                )
                StaffClubRelation.objects.create(staff=profile, club=club)

        def request_query_count():
            cache.clear()
            with CaptureQueriesContext(connection) as queries:
                response = self.client.get(
                    f"{reverse('clubs:staff_management')}?q=&page=1",
                    secure=True,
                )
            self.assertEqual(response.status_code, 200)
            return len(queries)

        self.client.force_login(self.admin)
        add_clubs(0, 5)
        small_dataset_queries = request_query_count()
        add_clubs(5, 25)
        large_dataset_queries = request_query_count()

        self.assertLessEqual(large_dataset_queries, small_dataset_queries)

    def test_user_dashboard_bulk_loads_clubs_channels_and_latest_submissions(self):
        profile = self.admin.profile
        profile.role = 'admin'
        profile.save(update_fields=['role'])
        channels = [
            FormChannel.objects.create(
                name=f'仪表板通道 {index}',
                slug=f'dashboard-channel-{index}',
                publish_status='published',
                submission_policy='once_total',
                is_active=True,
            )
            for index in range(5)
        ]
        def add_clubs(start, stop):
            clubs = []
            for index in range(start, stop):
                club = Club.objects.create(
                    name=f'仪表板社团 {index:02d}',
                    founded_date=date.today(),
                    members_count=20,
                )
                Officer.objects.create(
                    club=club,
                    user_profile=profile,
                    position='president',
                    appointed_date=date.today(),
                )
                StaffClubRelation.objects.create(staff=profile, club=club)
                clubs.append(club)
            FormSubmission.objects.bulk_create([
                FormSubmission(channel=channel, club=club, submitter=self.admin)
                for club in clubs
                for channel in channels
            ])

        def request_query_count():
            cache.clear()
            with CaptureQueriesContext(connection) as queries:
                response = self.client.get(reverse('clubs:user_dashboard'), secure=True)
            self.assertEqual(response.status_code, 200)
            return response, len(queries)

        self.client.force_login(self.admin)
        session = self.client.session
        session['active_identity'] = 'president'
        session.save()
        add_clubs(0, 5)
        _, small_dataset_queries = request_query_count()
        add_clubs(5, 20)
        response, large_dataset_queries = request_query_count()

        self.assertContains(response, '仪表板社团 19')
        self.assertLessEqual(large_dataset_queries, small_dataset_queries)

    def test_submission_display_title_uses_prefetched_values_without_queries(self):
        channel = FormChannel.objects.create(name='标题预取', slug='title-prefetch')
        field = FormField.objects.create(
            channel=channel,
            label='标题',
            field_key='title',
            field_type='text',
        )
        club = Club.objects.create(
            name='标题预取社团',
            founded_date=date.today(),
        )
        submission = FormSubmission.objects.create(
            channel=channel,
            club=club,
            submitter=self.admin,
        )
        FormFieldValue.objects.create(
            submission=submission,
            field=field,
            value_text='无需额外查询的标题',
        )
        loaded = FormSubmission.objects.prefetch_related('values__field').get(pk=submission.pk)

        with CaptureQueriesContext(connection) as queries:
            title = loaded.display_title

        self.assertEqual(title, '无需额外查询的标题')
        self.assertEqual(len(queries), 0)

    def test_my_room_bookings_is_paginated(self):
        room = Room.objects.create(name='分页测试房间')
        RoomBooking.objects.bulk_create([
            RoomBooking(
                room=room,
                user=self.admin,
                booking_date=date.today(),
                start_time=time(8, 0),
                end_time=time(9, 0),
                purpose=f'分页预约 {index:02d}',
                participant_count=1,
                contact_phone='13800000000',
            )
            for index in range(25)
        ])
        self.client.force_login(self.admin)

        response = self.client.get(f"{reverse('clubs:my_room_bookings')}?page=2", secure=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['bookings']), 1)
        self.assertContains(response, '预约记录分页')
        self.assertContains(response, '第 2 / 2 页，共 25 条')


class DynamicFormReReviewTests(TestCase):
    """已完成审核请求的重新审核与修改自己的审核结果。"""

    def setUp(self):
        self.admin = User.objects.create_superuser('re-admin', '', 'test-password')
        UserProfile.objects.get_or_create(
            user=self.admin,
            defaults={'role': 'admin', 'status': 'approved'},
        )
        self.reviewer = User.objects.create_user('re-reviewer', password='test-password')
        UserProfile.objects.create(user=self.reviewer, role='staff', status='approved')
        self.submitter = User.objects.create_user('re-submitter', password='test-password')
        UserProfile.objects.create(user=self.submitter, role='member', status='approved')
        self.club = Club.objects.create(name='重审社团', founded_date=date(2026, 1, 1))
        self.channel = FormChannel.objects.create(
            name='普通审核通道',
            slug='plain-rereview',
            is_active=True,
            publish_status='published',
            required_approval_count=2,
        )

    def make_rejected_submission(self, attempt=1):
        submission = FormSubmission.objects.create(
            channel=self.channel,
            club=self.club,
            submitter=self.submitter,
            status='rejected',
            resubmission_count=attempt,
            review_comment='原打回意见',
        )
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=self.reviewer,
            status='rejected',
            comment='原审核意见',
            submission_attempt=attempt,
        )
        return submission

    def test_admin_reenter_starts_new_review_round_keeping_history(self):
        submission = self.make_rejected_submission()
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'reenter', 'comment': '重新审核'},
            secure=True,
        )

        self.assertRedirects(
            response,
            reverse('clubs:staff_audit_center', args=[self.channel.slug]),
            fetch_redirect_response=False,
        )
        submission.refresh_from_db()
        self.assertEqual(submission.status, 'pending')
        self.assertEqual(submission.resubmission_count, 2)
        self.assertIsNone(submission.reviewer)
        self.assertIsNone(submission.reviewed_at)
        self.assertEqual(submission.review_comment, '')
        # 原审核记录保留在旧轮次，新轮次无任何投票
        self.assertEqual(submission.reviews.count(), 1)
        self.assertEqual(submission.reviews.first().submission_attempt, 1)
        self.assertFalse(submission.reviews.filter(submission_attempt=2).exists())
        self.assertEqual(submission.reviewer_count, 0)

    def test_admin_reenter_snapshots_previous_round_content_and_time(self):
        text_field = FormField.objects.create(
            channel=self.channel,
            field_key='reason',
            label='事由',
            field_type='text',
            is_active=True,
        )
        submission = FormSubmission.objects.create(
            channel=self.channel,
            club=self.club,
            submitter=self.submitter,
            status='rejected',
        )
        FormFieldValue.objects.create(
            submission=submission,
            field=text_field,
            value_text='第一轮内容',
            review_status='rejected',
        )
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=self.reviewer,
            status='rejected',
            comment='原审核意见',
            submission_attempt=1,
        )
        self.client.force_login(self.admin)

        self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'reenter', 'comment': '重新审核'},
            secure=True,
        )

        submission.refresh_from_db()
        history = submission.metadata.get('attempt_history') or []
        old_round = next(
            (entry for entry in history if entry.get('attempt') == 1),
            None,
        )
        self.assertIsNotNone(old_round)
        self.assertEqual(
            old_round['submitted_at'],
            timezone.localtime(submission.submitted_at).strftime('%Y-%m-%d %H:%M'),
        )
        self.assertEqual(old_round['status_label'], '已拒绝')
        self.assertEqual(old_round['fields'][0]['value'], '第一轮内容')

        # 新轮次收到投票后，时间线应同时展示两个轮次且都有提交时间
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=self.reviewer,
            status='approved',
            comment='第二轮通过',
            submission_attempt=2,
        )
        response = self.client.get(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            secure=True,
        )
        groups = response.context['attempt_groups']
        self.assertEqual(len(groups), 2)
        self.assertTrue(all(group.get('submitted_at') for group in groups))
        self.assertFalse(any(group.get('submitted_at_approx') for group in groups))

    def test_attempt_groups_do_not_fake_time_for_review_only_rounds(self):
        submission = FormSubmission.objects.create(
            channel=self.channel,
            club=self.club,
            submitter=self.submitter,
            status='pending',
            resubmission_count=2,
        )
        old_review = FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=self.reviewer,
            status='rejected',
            comment='第一轮意见',
            submission_attempt=1,
        )
        old_time = timezone.now() - timezone.timedelta(days=2)
        FormSubmissionReview.objects.filter(pk=old_review.pk).update(reviewed_at=old_time)
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=self.reviewer,
            status='approved',
            comment='第二轮意见',
            submission_attempt=2,
        )
        self.client.force_login(self.reviewer)

        response = self.client.get(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            secure=True,
        )
        groups = response.context['attempt_groups']
        self.assertEqual(len(groups), 2)
        first, second = groups
        # 无内容快照的旧轮次不伪造提交时间，交给模板显示“没有数据”
        self.assertNotIn('submitted_at', first)
        self.assertEqual(
            second['submitted_at'],
            timezone.localtime(submission.submitted_at).strftime('%Y-%m-%d %H:%M'),
        )
        self.assertNotIn('submitted_at_approx', second)

    def test_admin_reenter_approved_submission_allowed(self):
        submission = self.make_rejected_submission()
        submission.status = 'approved'
        submission.review_comment = '已通过'
        submission.save(update_fields=['status', 'review_comment'])
        FormSubmissionReview.objects.filter(submission=submission).update(status='approved')
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'reenter'},
            secure=True,
        )

        self.assertRedirects(
            response,
            reverse('clubs:staff_audit_center', args=[self.channel.slug]),
            fetch_redirect_response=False,
        )
        submission.refresh_from_db()
        self.assertEqual(submission.status, 'pending')
        self.assertEqual(submission.resubmission_count, 2)

    def test_non_admin_cannot_reenter(self):
        submission = self.make_rejected_submission()
        self.client.force_login(self.reviewer)

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'reenter'},
            secure=True,
        )
        followed = self.client.get(response.url, secure=True)
        self.assertContains(followed, '仅管理员可以重新审核')
        submission.refresh_from_db()
        self.assertEqual(submission.status, 'rejected')
        self.assertEqual(submission.resubmission_count, 1)

    def test_reenter_rejected_on_pending_submission(self):
        submission = FormSubmission.objects.create(
            channel=self.channel,
            club=self.club,
            submitter=self.submitter,
            status='pending',
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'reenter'},
            secure=True,
        )
        followed = self.client.get(response.url, secure=True)
        self.assertContains(followed, '仅已完成审核的请求可以重新审核')

    def test_business_action_channel_blocks_override_and_reenter(self):
        channel = FormChannel.objects.create(
            name='业务动作通道',
            slug='biz-rereview',
            builtin_action='club_application',
            is_active=True,
            publish_status='published',
        )
        submission = FormSubmission.objects.create(
            channel=channel,
            club=self.club,
            submitter=self.submitter,
            status='approved',
        )
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=self.reviewer,
            status='approved',
            submission_attempt=1,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'reenter'},
            secure=True,
        )
        self.assertContains(self.client.get(response.url, secure=True), '不允许重新审核')
        submission.refresh_from_db()
        self.assertEqual(submission.status, 'approved')

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'direct_reject'},
            secure=True,
        )
        self.assertContains(self.client.get(response.url, secure=True), '不允许覆盖修改')
        submission.refresh_from_db()
        self.assertEqual(submission.status, 'approved')

    def test_direct_modify_only_changes_own_review(self):
        second_reviewer = User.objects.create_user('re-reviewer-2', password='test-password')
        UserProfile.objects.create(user=second_reviewer, role='staff', status='approved')
        submission = self.make_rejected_submission()
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=second_reviewer,
            status='rejected',
            submission_attempt=1,
        )
        self.client.force_login(self.reviewer)

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'direct_approve', 'comment': '改为通过'},
            secure=True,
        )

        self.assertRedirects(
            response,
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            fetch_redirect_response=False,
        )
        submission.refresh_from_db()
        self.assertEqual(submission.status, 'approved')
        self.assertEqual(
            submission.reviews.get(reviewer=self.reviewer).status,
            'approved',
        )
        self.assertEqual(
            submission.reviews.get(reviewer=second_reviewer).status,
            'rejected',
        )
        # 审核人数不因修改而增加
        self.assertEqual(submission.reviewer_count, 2)
        self.assertEqual(submission.reviews.count(), 2)

    def test_reviewer_direct_modify_requires_participation(self):
        outsider = User.objects.create_user('re-outsider', password='test-password')
        UserProfile.objects.create(user=outsider, role='staff', status='approved')
        submission = self.make_rejected_submission()
        self.client.force_login(outsider)

        response = self.client.post(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            {'action': 'direct_approve'},
            secure=True,
        )
        self.assertContains(self.client.get(response.url, secure=True), '仅参与本次审核的审核人或管理员可以修改审核结果')
        submission.refresh_from_db()
        self.assertEqual(submission.status, 'rejected')

    def test_review_page_context_permissions(self):
        submission = self.make_rejected_submission()
        self.client.force_login(self.reviewer)
        response = self.client.get(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            secure=True,
        )
        self.assertTrue(response.context['can_edit_own_review'])
        self.assertFalse(response.context['can_reenter_review'])

        self.client.force_login(self.admin)
        response = self.client.get(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            secure=True,
        )
        self.assertTrue(response.context['can_edit_own_review'])
        self.assertTrue(response.context['can_reenter_review'])

    def test_revise_preserves_rejected_round_content_in_history(self):
        from django.core.files.base import ContentFile

        from ..models import FormUploadedFile
        from ..storage_backends import ClubStorage

        text_field = FormField.objects.create(
            channel=self.channel,
            field_key='reason',
            label='事由',
            field_type='text',
            is_active=True,
        )
        file_field = FormField.objects.create(
            channel=self.channel,
            field_key='attachment',
            label='附件',
            field_type='file',
            is_active=True,
        )
        self.submitter.profile.role = 'president'
        self.submitter.profile.save(update_fields=['role'])
        Officer.objects.create(
            club=self.club,
            user_profile=self.submitter.profile,
            position='president',
            appointed_date=date(2026, 1, 1),
        )

        from PIL import Image

        png_buffer = BytesIO()
        Image.new('RGB', (1, 1), 'white').save(png_buffer, format='PNG')
        png = png_buffer.getvalue()
        submission = FormSubmission.objects.create(
            channel=self.channel,
            club=self.club,
            submitter=self.submitter,
            status='rejected',
        )
        FormFieldValue.objects.create(
            submission=submission,
            field=text_field,
            value_text='第一轮内容',
            review_status='rejected',
        )
        old_upload = FormUploadedFile.objects.create(
            submission=submission,
            field=file_field,
            file=ContentFile(png, name='old.png'),
            original_name='旧附件.png',
            review_status='rejected',
        )
        old_pdf_upload = FormUploadedFile.objects.create(
            submission=submission,
            field=file_field,
            file=ContentFile(b'%PDF-1.4\nold', name='old.pdf'),
            original_name='旧附件.pdf',
            review_status='rejected',
        )
        old_storage_name = old_upload.file.name
        FormSubmissionReview.objects.create(
            submission=submission,
            reviewer=self.reviewer,
            status='rejected',
            comment='原审核意见',
            submission_attempt=1,
        )

        self.client.force_login(self.submitter)
        response = self.client.post(
            reverse('clubs:revise_dynamic_submission', args=[submission.public_id]),
            {
                f'field_{text_field.id}': '第二轮内容',
                f'field_{file_field.id}': SimpleUploadedFile('new.png', png, content_type='image/png'),
            },
            secure=True,
        )
        self.assertRedirects(
            response,
            reverse('clubs:approval_detail', args=[self.channel.slug, submission.public_id]),
            fetch_redirect_response=False,
        )

        submission.refresh_from_db()
        self.assertEqual(submission.resubmission_count, 2)
        self.assertEqual(submission.status, 'pending')
        # 旧文件记录已删除，但归档副本保留供历史下载
        self.assertEqual(submission.uploaded_files.count(), 1)

        history = submission.metadata.get('attempt_history') or []
        self.assertEqual(len(history), 1)
        entry = history[0]
        self.assertEqual(entry['attempt'], 1)
        self.assertEqual(entry['status'], 'rejected')
        self.assertTrue(
            any(
                field['value'] == '第一轮内容'
                for field in entry['fields']
                if field['field_key'] == 'reason'
            )
        )
        archived_files = [
            file for file in entry['files']
            if file['field_key'] == 'attachment'
        ]
        archived_file = archived_files[0]
        self.assertTrue(archived_file['storage_name'])
        self.assertNotEqual(archived_file['storage_name'], old_storage_name)
        self.assertTrue(ClubStorage().exists(archived_file['storage_name']))
        self.assertFalse(ClubStorage().exists(old_storage_name))
        self.assertEqual(len(archived_files), 2)
        pdf_file = next(
            file for file in archived_files
            if file['file_name'].endswith('.pdf')
        )
        self.assertFalse(pdf_file['is_image'])

        # 详情页展示历史内容
        detail = self.client.get(
            reverse('clubs:approval_detail', args=[self.channel.slug, submission.public_id]),
            secure=True,
        )
        self.assertContains(detail, '历史提交与审核记录')
        self.assertContains(detail, '第一轮内容')
        self.assertContains(detail, '旧附件.png')
        self.assertContains(detail, '旧附件.pdf')
        self.assertContains(detail, '原审核意见')
        self.assertContains(detail, 'open_in_new')
        self.assertContains(detail, '?inline=1')

        # 干事/管理员审核页同样展示历史内容
        self.client.force_login(self.admin)
        staff_page = self.client.get(
            reverse('clubs:staff_review_form_submission', args=[submission.public_id]),
            secure=True,
        )
        self.assertContains(staff_page, '历史提交与审核记录')
        self.assertContains(staff_page, '第一轮内容')

        # 历史附件仍可下载
        download = self.client.get(
            reverse(
                'clubs:history_submission_file',
                args=[submission.public_id, 1, 0],
            ),
            secure=True,
        )
        self.assertEqual(download.status_code, 200)
        self.assertEqual(b''.join(download.streaming_content), png)

        # 历史附件支持 inline 查看（非图片文件）
        pdf_index = next(
            index for index, file in enumerate(archived_files)
            if file['file_name'].endswith('.pdf')
        )
        inline_download = self.client.get(
            reverse(
                'clubs:history_submission_file',
                args=[submission.public_id, 1, pdf_index],
            )
            + '?inline=1',
            secure=True,
        )
        self.assertEqual(inline_download.status_code, 200)
        self.assertEqual(
            b''.join(inline_download.streaming_content),
            b'%PDF-1.4\nold',
        )


class ReviseSubmissionUploadTests(TestCase):
    """被打回提交的补交上传回归测试（含客户端 MD5、同内容去重与已注销审核人）。"""

    def setUp(self):
        from ..models import FormUploadedFile

        self.admin = User.objects.create_superuser(
            username='revise-upload-admin',
            email='',
            password='test-password',
        )
        UserProfile.objects.get_or_create(
            user=self.admin,
            defaults={'role': 'admin', 'status': 'approved'},
        )
        self.submitter = User.objects.create_user(
            'revise-upload-sub',
            password='test-password',
        )
        self.profile = UserProfile.objects.create(
            user=self.submitter,
            role='president',
            status='approved',
        )
        self.club = Club.objects.create(
            name='补交上传社团',
            founded_date=date(2026, 1, 1),
        )
        Officer.objects.create(
            club=self.club,
            user_profile=self.profile,
            position='president',
            appointed_date=date(2026, 1, 1),
        )
        self.channel = FormChannel.objects.create(
            name='补交通道',
            slug='revise-upload-channel',
            is_active=True,
        )
        self.field = FormField.objects.create(
            channel=self.channel,
            field_key='doc',
            label='附件',
            field_type='file',
            is_active=True,
        )
        self.submission = FormSubmission.objects.create(
            channel=self.channel,
            club=self.club,
            submitter=self.submitter,
            status='rejected',
        )
        FormUploadedFile.objects.create(
            submission=self.submission,
            field=self.field,
            file=SimpleUploadedFile(
                'old.pdf',
                b'%PDF-1.4 old',
                content_type='application/pdf',
            ),
            original_name='old.pdf',
            review_status='rejected',
        )
        self.client.force_login(self.submitter)

    def test_revise_upload_simple(self):
        new_file = SimpleUploadedFile(
            'new.pdf',
            b'%PDF-1.4 new',
            content_type='application/pdf',
        )
        response = self.client.post(
            reverse(
                'clubs:revise_dynamic_submission',
                args=[self.submission.public_id],
            ),
            {f'field_{self.field.id}': new_file},
            secure=True,
        )
        self.assertEqual(response.status_code, 302, response.content[:500])

    def test_revise_upload_with_client_md5(self):
        data = b'%PDF-1.4 md5-upload'
        new_file = SimpleUploadedFile(
            'new.pdf',
            data,
            content_type='application/pdf',
        )
        digest = md5(data).hexdigest()
        response = self.client.post(
            reverse(
                'clubs:revise_dynamic_submission',
                args=[self.submission.public_id],
            ),
            {
                f'field_{self.field.id}': new_file,
                f'md5_field_{self.field.id}': f'["{digest}"]',
            },
            secure=True,
        )
        self.assertEqual(response.status_code, 302, response.content.decode('utf-8', 'replace')[-3000:])

    def test_revise_upload_same_content_as_old(self):
        from ..models import FormUploadedFile

        data = b'%PDF-1.4 same-content'
        self.submission.uploaded_files.all().delete()
        FormUploadedFile.objects.create(
            submission=self.submission,
            field=self.field,
            file=SimpleUploadedFile(
                'old.pdf',
                data,
                content_type='application/pdf',
            ),
            original_name='old.pdf',
            review_status='rejected',
        )
        old_name = self.submission.uploaded_files.get().file.name
        new_file = SimpleUploadedFile(
            'new.pdf',
            data,
            content_type='application/pdf',
        )
        digest = md5(data).hexdigest()
        response = self.client.post(
            reverse(
                'clubs:revise_dynamic_submission',
                args=[self.submission.public_id],
            ),
            {
                f'field_{self.field.id}': new_file,
                f'md5_field_{self.field.id}': f'["{digest}"]',
            },
            secure=True,
        )
        self.assertEqual(response.status_code, 302, response.content[:500])
        self.submission.refresh_from_db()
        self.assertTrue(old_name)

    def test_revise_upload_with_missing_old_file(self):
        """旧文件物理缺失时补交不应 500（线上 404 坏图场景回归）。"""
        from ..models import FormUploadedFile

        self.submission.uploaded_files.all().delete()
        uploaded = FormUploadedFile.objects.create(
            submission=self.submission,
            field=self.field,
            file=SimpleUploadedFile(
                '图片.jpg',
                b'fake-old-image',
                content_type='image/jpeg',
            ),
            original_name='图片.jpg',
            review_status='rejected',
        )
        uploaded.file.storage.delete(uploaded.file.name)
        # 模拟线上内容寻址文件：blobs/ 路径但登记记录与物理文件都缺失
        uploaded.file.name = 'blobs/2cd8bde463f5d82aae0f0cec061d6b8f.jpg'
        uploaded.save(update_fields=['file'])

        new_file = SimpleUploadedFile(
            'new.png',
            base64.b64decode(
                'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=='
            ),
            content_type='image/png',
        )
        response = self.client.post(
            reverse(
                'clubs:revise_dynamic_submission',
                args=[self.submission.public_id],
            ),
            {f'field_{self.field.id}': new_file},
            secure=True,
        )
        from django.contrib.messages import get_messages

        debug_msgs = [str(item) for item in get_messages(response.wsgi_request)]
        self.assertEqual(
            response.status_code,
            302,
            f'messages={debug_msgs}',
        )

    def test_approval_detail_with_deleted_reviewer_does_not_500(self):
        from ..models import FormSubmissionReview

        reviewer = User.objects.create_user(
            'revise-upload-deleted-reviewer',
            password='test-password',
        )
        review = FormSubmissionReview.objects.create(
            submission=self.submission,
            reviewer=reviewer,
            status='rejected',
            comment='打回原因',
            submission_attempt=1,
        )
        reviewer.delete()
        review.refresh_from_db()
        self.assertIsNone(review.reviewer)

        response = self.client.get(
            reverse(
                'clubs:approval_detail',
                args=[self.channel.slug, self.submission.public_id],
            ),
            secure=True,
        )
        self.assertEqual(response.status_code, 200, response.content[:500])

    def test_revise_page_renders_revision_ui(self):
        response = self.client.get(
            reverse(
                'clubs:revise_dynamic_submission',
                args=[self.submission.public_id],
            ),
            secure=True,
        )
        self.assertEqual(response.status_code, 200, response.content[:500])
        self.assertContains(response, 'revision-banner')
        self.assertContains(response, 'revision-files')
        self.assertContains(response, 'revision-file-entry')
        self.assertContains(response, 'revision-replace-hint')
        self.assertContains(response, 'old.pdf')

    def test_revise_page_renders_rejected_image_inline(self):
        from ..models import FormUploadedFile

        self.submission.uploaded_files.all().delete()
        FormUploadedFile.objects.create(
            submission=self.submission,
            field=self.field,
            file=SimpleUploadedFile(
                'old.png',
                b'\x89PNG\r\n\x1a\nfake-image',
                content_type='image/png',
            ),
            original_name='old.png',
            review_status='rejected',
        )
        response = self.client.get(
            reverse(
                'clubs:revise_dynamic_submission',
                args=[self.submission.public_id],
            ),
            secure=True,
        )
        self.assertEqual(response.status_code, 200, response.content[:500])
        self.assertContains(response, 'revision-image-link')
        self.assertContains(response, '<img')
        self.assertContains(response, 'old.png')

    def test_revise_page_handles_missing_image_without_broken_img(self):
        from ..models import FormUploadedFile

        self.submission.uploaded_files.all().delete()
        uploaded = FormUploadedFile.objects.create(
            submission=self.submission,
            field=self.field,
            file=SimpleUploadedFile(
                'missing.png',
                b'\x89PNG\r\n\x1a\nfake-image',
                content_type='image/png',
            ),
            original_name='missing.png',
            review_status='rejected',
        )
        missing_url = uploaded.file.url
        uploaded.file.storage.delete(uploaded.file.name)

        response = self.client.get(
            reverse(
                'clubs:revise_dynamic_submission',
                args=[self.submission.public_id],
            ),
            secure=True,
        )
        self.assertEqual(response.status_code, 200, response.content[:500])
        self.assertContains(response, 'revision-file-missing')
        self.assertContains(response, '文件已丢失')
        self.assertNotContains(response, missing_url)


class DepartmentEditAndProfileRobustnessTests(TestCase):
    """审计修复回归测试：非法表单参数不产生 500，缺失 profile 的用户操作有兜底。"""

    def setUp(self):
        # 首启引导中间件要求系统存在管理员，否则所有请求都会重定向到 /oobe/
        self.admin = User.objects.create_superuser(
            username='robustness-admin',
            email='',
            password='test-password',
        )

    def _make_staff(self, username='dept-edit-staff'):
        staff = User.objects.create_user(username=username, password='test-password')
        UserProfile.objects.create(user=staff, role='staff', status='approved')
        self.client.force_login(staff)
        return staff

    def test_edit_department_rejects_non_numeric_order(self):
        from ..models import Department

        self._make_staff()
        dept = Department.objects.create(name='原部门', order=2)
        response = self.client.post(
            reverse('clubs:edit_department', args=[dept.pk]),
            {'name': '新部门', 'order': 'abc'},
            secure=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '排序顺序必须是大于或等于 0 的整数')
        dept.refresh_from_db()
        self.assertEqual(dept.name, '原部门')
        self.assertEqual(dept.order, 2)

    def test_edit_department_rejects_empty_name(self):
        from ..models import Department

        self._make_staff()
        dept = Department.objects.create(name='原部门')
        response = self.client.post(
            reverse('clubs:edit_department', args=[dept.pk]),
            {'name': '   ', 'order': '3'},
            secure=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '部门名称不能为空')
        dept.refresh_from_db()
        self.assertEqual(dept.name, '原部门')

    def test_edit_department_accepts_valid_update(self):
        from ..models import Department

        self._make_staff()
        dept = Department.objects.create(name='原部门', order=1)
        response = self.client.post(
            reverse('clubs:edit_department', args=[dept.pk]),
            {'name': '新部门', 'order': '5'},
            secure=True,
        )
        self.assertRedirects(
            response,
            reverse('clubs:manage_departments'),
            fetch_redirect_response=False,
        )
        dept.refresh_from_db()
        self.assertEqual(dept.name, '新部门')
        self.assertEqual(dept.order, 5)

    def test_edit_profile_without_profile_redirects(self):
        user = User.objects.create_user(username='no-profile-user', password='test-password')
        self.client.force_login(user)
        response = self.client.post(
            reverse('clubs:edit_profile'),
            {'action': 'update_info'},
            secure=True,
        )
        self.assertRedirects(
            response,
            reverse('clubs:index'),
            fetch_redirect_response=False,
        )
        self.assertTrue(User.objects.filter(pk=user.pk).exists())

    def test_delete_account_without_profile_redirects(self):
        user = User.objects.create_user(username='no-profile-del', password='test-password')
        self.client.force_login(user)
        response = self.client.post(
            reverse('clubs:delete_account'),
            {'confirm_username': user.username},
            secure=True,
        )
        self.assertRedirects(
            response,
            reverse('clubs:index'),
            fetch_redirect_response=False,
        )
        self.assertTrue(User.objects.filter(pk=user.pk).exists())

    def test_room_booking_str_without_profile(self):
        from ..models import Room, RoomBooking

        user = User.objects.create_user(username='booking-no-profile', password='test-password')
        room = Room.objects.create(name='测试房间', capacity=10)
        booking = RoomBooking.objects.create(
            room=room,
            user=user,
            booking_date=date(2026, 8, 6),
            start_time=time(9, 0),
            end_time=time(10, 0),
            purpose='测试',
            participant_count=2,
            contact_phone='13800000000',
        )
        self.assertIn('booking-no-profile', str(booking))


class FileBlobDedupTests(TestCase):
    """客户端 MD5 内容寻址去重：同文件只保存一份物理文件。"""

    @staticmethod
    def _make_file(name, data, content_type='application/octet-stream'):
        uploaded = SimpleUploadedFile(name, data, content_type=content_type)
        uploaded.client_md5 = md5(data).hexdigest()
        return uploaded

    def test_same_client_md5_reuses_single_blob_and_first_extension(self):
        data = b'same-content-for-dedup-test'
        first_name = default_storage.save(
            'uploads/a.jpg', self._make_file('a.jpg', data)
        )
        second_name = default_storage.save(
            'uploads/b.jpeg', self._make_file('b.jpeg', data)
        )

        self.assertEqual(first_name, second_name)
        self.assertTrue(first_name.startswith('blobs/'))
        self.assertTrue(first_name.endswith('.jpg'))
        self.assertTrue(default_storage.exists(first_name))

        blob = FileBlob.objects.get(md5=md5(data).hexdigest())
        self.assertEqual(blob.storage_name, first_name)
        self.assertEqual(blob.ref_count, 2)
        self.assertEqual(blob.size, len(data))
        self.assertEqual(FileBlob.objects.count(), 1)

        # 清理：释放两份引用后物理文件与登记记录都应消失
        default_storage.delete(first_name)
        self.assertTrue(default_storage.exists(first_name))
        default_storage.delete(second_name)
        self.assertFalse(default_storage.exists(first_name))
        self.assertFalse(FileBlob.objects.filter(md5=md5(data).hexdigest()).exists())

    def test_same_claimed_md5_different_size_falls_back_to_random_file(self):
        digest = md5(b'first-size').hexdigest()
        first_name = default_storage.save(
            'uploads/a.bin', self._make_file('a.bin', b'first-size')
        )
        # 客户端声称同一 MD5，但文件大小不同：不得复用旧 blob，必须单独落盘
        second = SimpleUploadedFile('b.bin', b'second-size-is-longer')
        second.client_md5 = digest
        second_name = default_storage.save('uploads/b.bin', second)

        self.assertNotEqual(first_name, second_name)
        self.assertTrue(first_name.startswith('blobs/'))
        self.assertFalse(second_name.startswith('blobs/'))
        with default_storage.open(second_name, 'rb') as handle:
            self.assertEqual(handle.read(), b'second-size-is-longer')

        blob = FileBlob.objects.get(md5=digest)
        self.assertEqual(blob.ref_count, 1)
        self.assertEqual(blob.size, len(b'first-size'))
        self.assertEqual(FileBlob.objects.count(), 1)

        default_storage.delete(first_name)
        default_storage.delete(second_name)
        self.assertFalse(default_storage.exists(first_name))
        self.assertFalse(FileBlob.objects.filter(md5=digest).exists())

    def test_submission_cascade_delete_releases_history_snapshot_reference(self):
        from ..models import FormUploadedFile
        from ..views.core import _snapshot_attempt_history

        data = b'history-snapshot-release-test'
        digest = md5(data).hexdigest()
        channel = FormChannel.objects.create(
            name='历史释放通道', slug='history-release', is_active=True
        )
        club = Club.objects.create(name='历史释放社团', founded_date=date(2026, 1, 1))
        submitter = User.objects.create_superuser(
            username='blob-history-admin', email='', password='test-password'
        )
        submission = FormSubmission.objects.create(
            channel=channel,
            club=club,
            submitter=submitter,
        )
        field = FormField.objects.create(
            channel=channel, field_key='doc', label='附件', field_type='file'
        )
        storage_name = default_storage.save(
            'form_submissions/x/doc.bin', self._make_file('doc.bin', data)
        )
        FormUploadedFile.objects.create(
            submission=submission,
            field=field,
            # 直接引用已存在的 blob 路径（字符串赋值不触发重写物理文件）
            file=storage_name,
            original_name='doc.bin',
        )

        _snapshot_attempt_history(submission)
        blob = FileBlob.objects.get(md5=digest)
        self.assertEqual(blob.ref_count, 2)

        # 直接级联删除提交（模拟通道/社团/用户删除），历史快照引用也必须释放
        submission.delete()

        self.assertFalse(FileBlob.objects.filter(md5=digest).exists())
        self.assertFalse(default_storage.exists(storage_name))

    def test_multiple_reenter_snapshots_release_all_references_on_delete(self):
        """多轮重新审核后整条删除：当前记录与每轮历史快照的引用都要释放。"""
        from ..models import FormUploadedFile
        from ..views.core import _snapshot_attempt_history

        data = b'multi-reenter-release-test'
        digest = md5(data).hexdigest()
        channel = FormChannel.objects.create(
            name='多轮重新审核通道', slug='multi-reenter', is_active=True
        )
        club = Club.objects.create(name='多轮重新审核社团', founded_date=date(2026, 1, 1))
        submitter = User.objects.create_superuser(
            username='blob-multi-reenter-admin', email='', password='test-password'
        )
        submission = FormSubmission.objects.create(
            channel=channel,
            club=club,
            submitter=submitter,
        )
        field = FormField.objects.create(
            channel=channel, field_key='doc', label='附件', field_type='file'
        )
        storage_name = default_storage.save(
            'form_submissions/x/doc.bin', self._make_file('doc.bin', data)
        )
        FormUploadedFile.objects.create(
            submission=submission,
            field=field,
            file=storage_name,
            original_name='doc.bin',
        )

        # 模拟两轮“重新审核”：每轮对当前文件快照 retain 一次
        _snapshot_attempt_history(submission)
        submission.resubmission_count += 1
        _snapshot_attempt_history(submission)
        submission.resubmission_count += 1
        submission.save()

        blob = FileBlob.objects.get(md5=digest)
        # 当前记录 1 份 + 两轮历史快照各 1 份
        self.assertEqual(blob.ref_count, 3)

        submission.delete()

        self.assertFalse(FileBlob.objects.filter(md5=digest).exists())
        self.assertFalse(default_storage.exists(storage_name))

    def test_field_example_file_replace_and_clear_release_old_blob(self):
        admin = User.objects.create_superuser(
            username='blob-field-admin', email='', password='test-password'
        )
        self.client.force_login(admin)
        channel = FormChannel.objects.create(
            name='示例替换通道', slug='example-replace', is_active=True
        )
        field = FormField.objects.create(
            channel=channel, field_key='doc', label='示例', field_type='file'
        )

        first_data = b'%PDF-1.4 example-first'
        first_digest = md5(first_data).hexdigest()
        response = self.client.post(
            reverse('clubs:edit_form_field', args=[channel.id, field.id]),
            {
                'label': '示例',
                'field_key': 'doc',
                'field_type': 'file',
                'md5_example_file': json.dumps([first_digest]),
                'example_file': SimpleUploadedFile(
                    'example.pdf', first_data, content_type='application/pdf'
                ),
            },
        )
        self.assertEqual(response.status_code, 302)
        first_blob = FileBlob.objects.get(md5=first_digest)
        self.assertEqual(first_blob.ref_count, 1)
        field.refresh_from_db()
        self.assertEqual(field.example_file.name, first_blob.storage_name)

        # 替换示例文件：旧 blob 引用必须释放并物理删除
        second_data = b'%PDF-1.4 example-second-longer'
        second_digest = md5(second_data).hexdigest()
        response = self.client.post(
            reverse('clubs:edit_form_field', args=[channel.id, field.id]),
            {
                'label': '示例',
                'field_key': 'doc',
                'field_type': 'file',
                'md5_example_file': json.dumps([second_digest]),
                'example_file': SimpleUploadedFile(
                    'example.pdf', second_data, content_type='application/pdf'
                ),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(FileBlob.objects.filter(md5=first_digest).exists())
        self.assertFalse(default_storage.exists(first_blob.storage_name))
        second_blob = FileBlob.objects.get(md5=second_digest)
        self.assertEqual(second_blob.ref_count, 1)
        field.refresh_from_db()
        self.assertEqual(field.example_file.name, second_blob.storage_name)

        # 清空示例文件：引用必须释放并物理删除
        response = self.client.post(
            reverse('clubs:edit_form_field', args=[channel.id, field.id]),
            {
                'label': '示例',
                'field_key': 'doc',
                'field_type': 'file',
                'clear_example': 'on',
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(FileBlob.objects.filter(md5=second_digest).exists())
        self.assertFalse(default_storage.exists(second_blob.storage_name))
        field.refresh_from_db()
        self.assertFalse(field.example_file)

    def test_delete_decrements_ref_count_before_last_reference(self):
        data = b'ref-count-release-test'
        digest = md5(data).hexdigest()
        first_name = default_storage.save(
            'uploads/first.bin', self._make_file('first.bin', data)
        )
        second_name = default_storage.save(
            'uploads/second.bin', self._make_file('second.bin', data)
        )
        self.assertEqual(first_name, second_name)

        default_storage.delete(first_name)
        blob = FileBlob.objects.get(md5=digest)
        self.assertEqual(blob.ref_count, 1)
        self.assertTrue(default_storage.exists(first_name))

        default_storage.delete(second_name)
        self.assertFalse(default_storage.exists(first_name))
        self.assertFalse(FileBlob.objects.filter(md5=digest).exists())

    def test_retain_protects_historical_snapshot(self):
        data = b'retain-snapshot-test'
        digest = md5(data).hexdigest()
        name = default_storage.save(
            'uploads/snapshot.bin', self._make_file('snapshot.bin', data)
        )

        self.assertTrue(default_storage.retain(name))
        blob = FileBlob.objects.get(md5=digest)
        self.assertEqual(blob.ref_count, 2)

        # 释放原引用后文件仍被历史快照保留
        default_storage.delete(name)
        self.assertTrue(default_storage.exists(name))
        self.assertEqual(FileBlob.objects.get(md5=digest).ref_count, 1)

        # 历史快照也释放后物理文件删除
        default_storage.delete(name)
        self.assertFalse(default_storage.exists(name))
        self.assertFalse(FileBlob.objects.filter(md5=digest).exists())

    def test_avatar_path_skips_dedup_even_with_client_md5(self):
        data = b'avatar-content-not-deduped'
        uploaded = self._make_file('avatar.png', data)
        name = default_storage.save('avatars/avatar.png', uploaded)

        self.assertTrue(name.startswith('avatars/'))
        self.assertFalse(FileBlob.objects.exists())
        default_storage.delete(name)
        self.assertFalse(default_storage.exists(name))

    def test_missing_fileblob_table_query_failure_falls_back_to_random_name(self):
        """FileBlob 表缺失（迁移未应用）时查询失败必须降级，不能 500。"""
        from django.db.utils import ProgrammingError

        data = b'missing-table-query-fallback'
        uploaded = self._make_file('missing.bin', data)
        with patch.object(
            FileBlob.objects,
            'filter',
            side_effect=ProgrammingError('no such table: clubs_fileblob'),
        ):
            name = default_storage.save('uploads/missing.bin', uploaded)

        self.assertFalse(name.startswith('blobs/'))
        with default_storage.open(name, 'rb') as handle:
            self.assertEqual(handle.read(), data)
        default_storage.delete(name)
        self.assertFalse(default_storage.exists(name))

    def test_missing_fileblob_table_create_failure_falls_back_and_cleans_blob(self):
        """登记 FileBlob 失败时已写出的 blobs 物理文件必须清理并降级保存。"""
        from django.db.utils import ProgrammingError

        data = b'missing-table-create-fallback'
        uploaded = self._make_file('missing.bin', data)
        with patch.object(
            FileBlob.objects,
            'filter',
            return_value=FileBlob.objects.none(),
        ), patch.object(
            FileBlob.objects,
            'create',
            side_effect=ProgrammingError('no such table: clubs_fileblob'),
        ):
            name = default_storage.save('uploads/missing.bin', uploaded)

        self.assertFalse(name.startswith('blobs/'))
        with default_storage.open(name, 'rb') as handle:
            self.assertEqual(handle.read(), data)
        default_storage.delete(name)
        self.assertFalse(default_storage.exists(name))

    def test_bind_client_md5_binds_in_order_and_ignores_invalid(self):
        post = QueryDict(mutable=True)
        post['md5_example_files'] = json.dumps([
            md5(b'first').hexdigest(),
            'not-a-valid-md5',
        ])
        first = SimpleUploadedFile('first.bin', b'first')
        second = SimpleUploadedFile('second.bin', b'second')
        files = MultiValueDict({'example_files': [first, second]})

        bind_client_md5_from_post(post, files)

        self.assertEqual(first.client_md5, md5(b'first').hexdigest())
        self.assertFalse(hasattr(second, 'client_md5'))

    def test_bind_client_md5_degrades_gracefully_on_bad_json(self):
        post = QueryDict(mutable=True)
        post['md5_example_files'] = '{bad json'
        uploaded = SimpleUploadedFile('file.bin', b'data')
        files = MultiValueDict({'example_files': [uploaded]})

        bind_client_md5_from_post(post, files)

        self.assertFalse(hasattr(uploaded, 'client_md5'))
